#!/usr/bin/env python3
"""Cassette structure discovery from accepted pieces."""

from __future__ import annotations

import argparse
import colorsys
import csv
import hashlib
import json
import re
import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path

_PKG_DIR = Path(__file__).resolve().parent
if str(_PKG_DIR) not in sys.path:
    sys.path.insert(0, str(_PKG_DIR))

from flank_interval import (
    describe as describe_intervals,
    extend_to_flanks,
    genes_in_span,
    augment as augment_with_span,
    orient,
    reference_rank,
)
from locus_scale import (
    DERIVE,
    describe as describe_window,
    reference_length,
    resolve as resolve_window,
)
from writers import (
    IDENTITY_HEADERS,
    IDENTITY_WIDTHS,
    STATUS_TINT,
    diagnostics_dir,
    find_output,
    freeze_after_status,
    tables_dir,
)

try:
    from .config_utils import (get_target_cfg, load_locus_cfg, load_settings,
                               resolve_target_name)
except ImportError:
    from config_utils import (get_target_cfg, load_locus_cfg, load_settings,
                              resolve_target_name)


MEMBER_ZONE = "cluster"
FLANK_ZONES = {"flank_left", "flank_right"}
CONTEXT_ZONES = {"cluster", "flank_left", "flank_right"}


def _read_csv(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(f"required LocusRuler output not found: {path}")
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = []
        seen = set()
        for row in rows:
            for key in row:
                if key not in seen:
                    fieldnames.append(key)
                    seen.add(key)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _integer(value, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _piece_index(value) -> str:
    return str(value if value is not None else "").strip()


def _piece_sort_key(piece: dict) -> tuple:
    return (
        _integer(piece.get("q_start"), 10**15),
        _integer(piece.get("q_end"), 10**15),
        _piece_index(piece.get("piece_idx")),
    )


def _gene_coordinate_key(gene: dict, piece: dict) -> tuple:
    start = _integer(gene.get("gene_start"))
    end = _integer(gene.get("gene_end"))
    if str(piece.get("strand", "+")) == "-":
        return (-max(start, end), -min(start, end), str(gene.get("locus_tag", "")))
    return (min(start, end), max(start, end), str(gene.get("locus_tag", "")))


def _family_token(gene: dict) -> str:
    family = str(gene.get("family") or "").strip()
    return family if family else "unassigned"


def _annotation_token(gene: dict) -> str:
    family = str(gene.get("family") or "").strip()
    if family:
        return family
    product = re.sub(r"[^a-z0-9]+", "_", str(gene.get("product") or "").casefold())
    product = product.strip("_") or "unknown"
    return f"unassigned[{product[:48]}]"


def _structure_id(signature: str) -> str:
    if signature in {"", "absent", "empty"}:
        return ""
    digest = hashlib.sha256(signature.encode("utf-8")).hexdigest()[:12]
    return f"CS_{digest}"


def _state_token(gene: dict) -> str:
    state = str(gene.get("state") or "").strip().upper() or "UNASSIGNED"
    return f"{_family_token(gene)}:{state}"


def _alignments(genes: list[dict]) -> set[tuple[str, str]]:
    """Distinct anchor alignments behind a run of genes.

    Two neighbours that carry the same coverage and identity are two CDS the
    one alignment ran across -- a gene the annotation split, not a duplication.
    Genuine copies each align on their own and disagree on at least one of the
    two. This is what separates a split gene from a tandem pair when both CDS
    are annotated complete and so both read as INTACT.
    """
    recorded = set()
    for gene in genes:
        coverage = str(gene.get("tblastn_cov") or "").strip()
        identity = str(gene.get("tblastn_pid") or "").strip()
        if coverage and identity:
            recorded.add((coverage, identity))
    return recorded


def _signature_members(ordered_genes: list[dict]) -> list[dict]:
    """Collapse adjacent fragments of one family without hiding true copies."""
    members = [gene for gene in ordered_genes if gene["is_cassette_member"] == "Y"]
    collapsed: list[dict] = []
    index = 0
    while index < len(members):
        family = _family_token(members[index])
        run_end = index + 1
        while (
            run_end < len(members)
            and _family_token(members[run_end]) == family
            and members[run_end]["piece_order"] == members[index]["piece_order"]
        ):
            run_end += 1
        run = members[index:run_end]
        if family == "unassigned":
            collapsed.extend(run)
        else:
            intact = [
                gene for gene in run
                if str(gene.get("state") or "").strip().upper() == "INTACT"
            ]
            # one shared alignment means one gene the annotation split; anything
            # else (distinct alignments, or none recorded) stays as copies
            if len(intact) >= 2 and len(_alignments(intact)) != 1:
                collapsed.extend(intact)
            else:
                collapsed.append(max(
                    run,
                    key=lambda gene: (
                        str(gene.get("state") or "").strip().upper() == "INTACT",
                        _integer(gene.get("gene_end")) - _integer(gene.get("gene_start")),
                    ),
                ))
        index = run_end
    return collapsed


def _accepted_pieces(rows: list[dict], locus_id: str) -> dict[str, dict[str, dict]]:
    by_genome: dict[str, dict[str, dict]] = defaultdict(dict)
    for row in rows:
        if row.get("locus_id") != locus_id:
            continue
        idx = _piece_index(row.get("piece_idx"))
        if idx:
            by_genome[row["genome_acc"]][idx] = row
    return dict(by_genome)


# Co-location window for cassette membership, as a multiple of the reference locus.
CASSETTE_SPAN_MULTIPLE = 5.0

# Identity columns from loci.xlsx, then this view's columns, then the gene grid.
MATRIX_FIXED_HEADERS = IDENTITY_HEADERS + [
    "locus_status",
    "structure_id",
    "structure_signature",
    "n_cassette_genes",
]
MATRIX_FIXED_WIDTHS = {**IDENTITY_WIDTHS, "locus_status": 16,
                       "structure_id": 18, "structure_signature": 55,
                       "n_cassette_genes": 12}


def _piece_span(piece: dict) -> tuple[str, int, int]:
    lo = _integer(piece.get("s_lo"))
    hi = _integer(piece.get("s_hi"))
    return str(piece.get("contig") or ""), min(lo, hi), max(lo, hi)


def _colocated_pieces(
    pieces: dict[str, dict],
    max_span_bp: int,
) -> tuple[dict[str, dict], list[str]]:
    """Split accepted pieces into cassette-colocated and off-locus."""
    if len(pieces) <= 1:
        return dict(pieces), []

    def _cover(p: dict) -> tuple[int, int]:
        return _integer(p.get("n_covered_genes")), _integer(p.get("aligned_bp"))

    body_idx = max(pieces, key=lambda i: _cover(pieces[i]))
    body = pieces[body_idx]
    b_contig, b_lo, b_hi = _piece_span(body)
    b_edge = str(body.get("contig_edge_proximal") or "").strip().upper() == "Y"

    kept: dict[str, dict] = {body_idx: body}
    dropped: list[str] = []
    for idx, piece in pieces.items():
        if idx == body_idx:
            continue
        p_contig, p_lo, p_hi = _piece_span(piece)
        p_edge = str(piece.get("contig_edge_proximal") or "").strip().upper() == "Y"
        if p_contig and p_contig == b_contig:
            gap = max(p_lo - b_hi, b_lo - p_hi, 0)
            if gap <= max_span_bp:
                kept[idx] = piece
                continue
        elif b_edge and p_edge:
            kept[idx] = piece
            continue
        dropped.append(idx)
    return kept, dropped


def _renumber(genes: list[dict], ref_rank: dict[str, int]) -> list[dict]:
    """Renumber after the extension added rows, keeping flanks on the outside."""
    left = [g for g in genes if g.get("zone") == "flank_left"]
    right = [g for g in genes if g.get("zone") == "flank_right"]
    middle = [g for g in genes if g.get("zone") not in FLANK_ZONES]

    def _start(gene):
        try:
            return int(gene.get("gene_start") or 0)
        except (TypeError, ValueError):
            return 0

    ordered: list[dict] = []
    for gene in (sorted(left, key=_start)
                 + orient(middle, ref_rank)
                 + sorted(right, key=_start)):
        gene = dict(gene)
        member = gene.get("zone") == MEMBER_ZONE
        gene["structure_order"] = len(ordered) + 1
        gene.setdefault("piece_order", 1)
        gene["within_piece_order"] = len(ordered) + 1
        gene["is_cassette_member"] = "Y" if member else ""
        gene.setdefault(
            "membership_source",
            "accepted_locus_piece" if member else "flank_context",
        )
        ordered.append(gene)
    return ordered


def _ordered_genes(
    diagnostic_rows: list[dict],
    pieces: dict[str, dict],
) -> list[dict]:
    """Return diagnostic genes on accepted pieces in reference-query order."""
    by_piece: dict[str, list[dict]] = defaultdict(list)
    seen = set()
    for row in diagnostic_rows:
        idx = _piece_index(row.get("piece_idx"))
        if idx not in pieces or row.get("zone") not in CONTEXT_ZONES:
            continue
        key = (
            idx,
            row.get("gene_contig"),
            row.get("gene_start"),
            row.get("gene_end"),
            row.get("locus_tag"),
            row.get("zone"),
        )
        if key in seen:
            continue
        seen.add(key)
        by_piece[idx].append(row)

    ordered: list[dict] = []
    for piece_order, piece in enumerate(sorted(pieces.values(), key=_piece_sort_key), start=1):
        idx = _piece_index(piece.get("piece_idx"))
        genes = sorted(by_piece.get(idx, []), key=lambda gene: _gene_coordinate_key(gene, piece))
        for within_piece_order, source in enumerate(genes, start=1):
            gene = dict(source)
            gene["piece_order"] = piece_order
            gene["within_piece_order"] = within_piece_order
            gene["structure_order"] = len(ordered) + 1
            gene["is_cassette_member"] = "Y" if gene.get("zone") == MEMBER_ZONE else ""
            gene["membership_source"] = "accepted_locus_piece"
            gene["piece_q_start"] = piece.get("q_start", "")
            gene["piece_q_end"] = piece.get("q_end", "")
            gene["piece_strand"] = piece.get("strand", "")
            ordered.append(gene)
    return ordered


def _signatures(ordered_genes: list[dict]) -> tuple[str, str, str, str]:
    members = _signature_members(ordered_genes)
    if not members:
        return "empty", "empty", "empty", ""

    family_tokens = [_family_token(gene) for gene in members]
    annotation_tokens = [_annotation_token(gene) for gene in members]
    state_tokens = [_state_token(gene) for gene in members]

    assembly_tokens: list[str] = []
    last_piece = None
    for gene, token in zip(members, family_tokens):
        piece = gene["piece_order"]
        if last_piece is not None and piece != last_piece:
            assembly_tokens.append("||")
        assembly_tokens.append(token)
        last_piece = piece
    return (
        ">".join(family_tokens),
        ">".join(annotation_tokens),
        ">".join(state_tokens),
        ">".join(assembly_tokens),
    )


def _family_color(family: str) -> tuple[float, float, float]:
    if family == "unassigned":
        return (0.78, 0.78, 0.78)
    digest = hashlib.sha256(family.encode("utf-8")).digest()
    hue = int.from_bytes(digest[:2], "big") / 65535
    return colorsys.hsv_to_rgb(hue, 0.58, 0.78)


def _hex_color(color) -> str:
    if isinstance(color, str):
        value = color.lstrip("#")
        if len(value) in {6, 8}:
            return value.upper()
    if isinstance(color, (tuple, list)) and len(color) >= 3:
        return "".join(
            f"{max(0, min(255, round(float(component) * 255))):02X}"
            for component in color[:3]
        )
    raise ValueError(f"unsupported color value: {color!r}")


def _display_key(gene: dict) -> str:
    family = str(gene.get("family") or "").strip()
    if family and family != "unassigned":
        return family
    label = str(gene.get("label") or "").strip()
    if label.startswith("[") and label.endswith("]"):
        label = label[1:-1].strip()
    return label if label and label != "unassigned" else "unassigned"


def _display_genes(ordered_genes: list[dict]) -> list[dict]:
    """Return cassette members plus defensible outer flank context."""
    member_pieces = {
        gene["piece_order"]
        for gene in ordered_genes
        if gene.get("is_cassette_member") == "Y"
    }
    single_piece = len(member_pieces) == 1
    first_piece = min(member_pieces) if member_pieces else None
    last_piece = max(member_pieces) if member_pieces else None
    return [
        gene for gene in ordered_genes
        if (
            gene.get("is_cassette_member") == "Y"
            or (
                gene.get("zone") in {"flank_left", "flank_right"}
                and (
                    single_piece
                    or (
                        gene.get("zone") == "flank_left"
                        and gene.get("piece_order") == first_piece
                    )
                    or (
                        gene.get("zone") == "flank_right"
                        and gene.get("piece_order") == last_piece
                    )
                )
                and (
                    single_piece
                    or _display_key(gene) != "unassigned"
                )
            )
        )
    ]


def _load_gene_strands(db_path: Path, diagnostic_rows: list[dict]) -> dict[tuple[str, str], str]:
    """Load strand metadata without embedding locus-specific assumptions."""
    locus_tags = sorted({
        row.get("locus_tag", "")
        for row in diagnostic_rows
        if row.get("locus_tag")
    })
    if not locus_tags or not db_path.exists():
        return {}
    found: dict[tuple[str, str], str] = {}
    with sqlite3.connect(db_path) as conn:
        for offset in range(0, len(locus_tags), 800):
            chunk = locus_tags[offset:offset + 800]
            placeholders = ",".join("?" for _ in chunk)
            query = (
                "SELECT genome_acc, locus_tag, strand FROM proteins "
                f"WHERE locus_tag IN ({placeholders})"
            )
            for genome_acc, locus_tag, strand in conn.execute(query, chunk):
                found[(str(genome_acc), str(locus_tag))] = str(strand or "")
    return found


def _display_settings(locus_cfg: dict) -> dict:
    configured = locus_cfg.get("cassette_display", {})
    if not isinstance(configured, dict):
        raise ValueError("cassette_display must be a JSON object")
    family_styles = configured.get("family_styles", {})
    if not isinstance(family_styles, dict):
        raise ValueError("cassette_display.family_styles must be a JSON object")
    product_styles = configured.get("product_styles", [])
    if not isinstance(product_styles, list):
        raise ValueError("cassette_display.product_styles must be a JSON array")
    for rule in product_styles:
        if (
            not isinstance(rule, dict)
            or not str(rule.get("pattern") or "")
            or not str(rule.get("style") or "")
        ):
            raise ValueError(
                "each cassette_display.product_styles entry requires "
                "pattern and style strings"
            )
        re.compile(str(rule["pattern"]))
    return {
        "scale_bp": float(configured.get("scale_bp", 620.0)),
        "scale_bar_bp": int(configured.get("scale_bar_bp", 2000)),
        "piece_gap_bp": int(configured.get("piece_gap_bp", 350)),
        "edge_color": str(configured.get("edge_color", "#2C3E50")),
        "context_color": str(configured.get("context_color", "#D7DBDD")),
        "core_connector_color": str(
            configured.get("core_connector_color", "#9FC3B5")
        ),
        "flank_connector_color": str(
            configured.get("flank_connector_color", "#B8A88A")
        ),
        "show_unconfigured_labels": bool(
            configured.get("show_unconfigured_labels", True)
        ),
        "family_styles": family_styles,
        "product_styles": product_styles,
    }


def _style_for_gene(gene: dict, display: dict) -> dict:
    key = _display_key(gene)
    if key == "unassigned" and gene.get("zone") in {"flank_left", "flank_right"}:
        product = str(gene.get("product") or "")
        for rule in display["product_styles"]:
            if re.search(str(rule["pattern"]), product):
                key = str(rule["style"])
                break
    configured = display["family_styles"].get(key, {})
    if not isinstance(configured, dict):
        raise ValueError(
            f"cassette_display.family_styles.{key} must be a JSON object"
        )
    if "color" in configured:
        color = configured["color"]
    elif gene.get("zone") in {"flank_left", "flank_right"}:
        color = display["context_color"]
    else:
        color = _family_color(key)
    label = configured.get("label")
    if label is None:
        label = (
            key
            if (
                key != "unassigned"
                and gene.get("zone") not in {"flank_left", "flank_right"}
                and display["show_unconfigured_labels"]
            )
            else ""
        )
    return {
        "key": key,
        "color": color,
        "label": str(label),
        "italic": bool(configured.get("italic", bool(label))),
    }


def _display_strand(gene: dict) -> str:
    strand = str(gene.get("strand") or "+")
    if str(gene.get("piece_strand") or "+") == "-":
        return "-" if strand != "-" else "+"
    return strand


def _gene_length(gene: dict) -> int:
    return max(
        abs(_integer(gene.get("gene_end")) - _integer(gene.get("gene_start"))) + 1,
        1,
    )


# Visual cap on any single gap, in gene-width units. A piece boundary or a
# stray far-flung gene would otherwise stretch one row's real bp distance
# across the whole shared axis and squeeze every other row to a sliver; past
# this cap the gap is drawn as a dashed break instead of to scale.
_MAX_GAP_UNITS = 1.5


def _layout_row(
    genes: list[dict],
    display: dict,
) -> tuple[list[tuple[dict, float, float, bool]], float]:
    scale_bp = display["scale_bp"]
    laid_out: list[tuple[dict, float, float, bool]] = []
    x = 0.0
    previous = None
    for gene in genes:
        break_before = False
        if previous is not None:
            piece_changed = gene["piece_order"] != previous["piece_order"]
            if piece_changed:
                gap_bp = display["piece_gap_bp"]
            else:
                prev_lo = min(
                    _integer(previous.get("gene_start")),
                    _integer(previous.get("gene_end")),
                )
                prev_hi = max(
                    _integer(previous.get("gene_start")),
                    _integer(previous.get("gene_end")),
                )
                gene_lo = min(
                    _integer(gene.get("gene_start")),
                    _integer(gene.get("gene_end")),
                )
                gene_hi = max(
                    _integer(gene.get("gene_start")),
                    _integer(gene.get("gene_end")),
                )
                gap_bp = max(
                    gene_lo - prev_hi - 1,
                    prev_lo - gene_hi - 1,
                    0,
                )
            raw_units = gap_bp / scale_bp
            break_before = piece_changed or raw_units > _MAX_GAP_UNITS
            x += min(raw_units, _MAX_GAP_UNITS)
        width = max(_gene_length(gene) / scale_bp, 0.14)
        laid_out.append((gene, x, width, break_before))
        x += width
        previous = gene
    return laid_out, x


def _draw_gene(
    ax,
    x: float,
    width: float,
    y: float,
    strand: str,
    color,
    edge_color: str,
) -> None:
    from matplotlib.patches import Polygon

    height = 0.54
    head = min(max(width * 0.28, 0.08), 0.30)
    if strand == "-":
        points = [
            (x + width, y - height / 2),
            (x + head, y - height / 2),
            (x, y),
            (x + head, y + height / 2),
            (x + width, y + height / 2),
        ]
    else:
        points = [
            (x, y - height / 2),
            (x + width - head, y - height / 2),
            (x + width, y),
            (x + width - head, y + height / 2),
            (x, y + height / 2),
        ]
    ax.add_patch(Polygon(
        points,
        closed=True,
        facecolor=color,
        edgecolor=edge_color,
        linewidth=0.8,
    ))


def _plot_catalog(
    catalog_rows: list[dict],
    representative_genes: dict[str, list[dict]],
    out_path: Path,
    display: dict,
) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams["font.family"] = "Arial"
    plt.rcParams["font.sans-serif"] = [
        "Arial",
        "Liberation Sans",
        "DejaVu Sans",
    ]
    shown = catalog_rows
    if not shown:
        return
    fig, ax = plt.subplots(figsize=(15, max(3.0, len(shown) * 0.82 + 1.5)))
    max_x = 1.0
    row_centers: list[tuple[dict[str, float], dict[str, str], float]] = []
    for row_index, catalog in enumerate(shown):
        y = len(shown) - row_index - 1
        structure_id = catalog["structure_id"]
        genes = _display_genes(representative_genes.get(structure_id, []))
        laid_out, row_width = _layout_row(genes, display)
        label_widths: dict[str, float] = {}
        for gene, _, width, _ in laid_out:
            key = _style_for_gene(gene, display)["key"]
            label_widths[key] = max(label_widths.get(key, 0.0), width)
        labeled: set[str] = set()
        centers: dict[str, float] = {}
        zones: dict[str, str] = {}
        prev_end = None
        for gene, x, width, break_before in laid_out:
            if break_before and prev_end is not None:
                ax.plot(
                    [prev_end, x],
                    [y, y],
                    color="#78909C",
                    linewidth=1.0,
                    linestyle="--",
                )
            style = _style_for_gene(gene, display)
            _draw_gene(
                ax,
                x,
                width,
                y,
                _display_strand(gene),
                style["color"],
                display["edge_color"],
            )
            if (
                style["label"]
                and width >= 0.38
                and width == label_widths[style["key"]]
                and style["key"] not in labeled
            ):
                label = (
                    style["label"]
                    if len(style["label"]) <= 18
                    else style["label"][:17] + "…"
                )
                ax.text(
                    x + width / 2,
                    y + 0.39,
                    label,
                    ha="center",
                    va="bottom",
                    fontsize=7.0,
                    fontstyle="italic" if style["italic"] else "normal",
                )
                labeled.add(style["key"])
            centers[style["key"]] = x + width / 2
            zones[style["key"]] = str(gene.get("zone") or "")
            prev_end = x + width
        row_centers.append((centers, zones, y))
        max_x = max(max_x, row_width)
        species = str(catalog.get("representative_species") or "")
        strain = str(catalog.get("representative_strain") or "")
        example = " ".join(
            part for part in (species, strain) if part
        ).strip()
        ax.text(
            -0.18,
            y,
            f"{structure_id} (n={catalog['n_genomes']})"
            + (f"\n{example}" if example else ""),
            ha="right",
            va="center",
            fontsize=8,
        )

    for (top, top_zones, top_y), (bottom, bottom_zones, bottom_y) in zip(
        row_centers[:-1],
        row_centers[1:],
    ):
        for key in set(top) & set(bottom):
            flank = (
                top_zones.get(key) in {"flank_left", "flank_right"}
                or bottom_zones.get(key) in {"flank_left", "flank_right"}
            )
            ax.plot(
                [top[key], bottom[key]],
                [top_y - 0.30, bottom_y + 0.30],
                color=(
                    display["flank_connector_color"]
                    if flank
                    else display["core_connector_color"]
                ),
                linewidth=0.8,
                alpha=0.45,
                zorder=0,
            )

    bar_width = display["scale_bar_bp"] / display["scale_bp"]
    bar_x = max(0.0, max_x - bar_width)
    bar_y = -0.58
    ax.plot(
        [bar_x, bar_x + bar_width],
        [bar_y, bar_y],
        color="#222222",
        linewidth=1.2,
    )
    ax.plot(
        [bar_x, bar_x],
        [bar_y - 0.04, bar_y + 0.04],
        color="#222222",
        linewidth=1.2,
    )
    ax.plot(
        [bar_x + bar_width, bar_x + bar_width],
        [bar_y - 0.04, bar_y + 0.04],
        color="#222222",
        linewidth=1.2,
    )
    scale_label = (
        f"{display['scale_bar_bp'] // 1000} kb"
        if display["scale_bar_bp"] % 1000 == 0
        else f"{display['scale_bar_bp']} bp"
    )
    ax.text(
        bar_x + bar_width / 2,
        bar_y - 0.10,
        scale_label,
        ha="center",
        va="top",
        fontsize=8,
    )
    ax.set_xlim(-5.8, max_x + 0.4)
    ax.set_ylim(-0.85, len(shown) - 0.15)
    ax.axis("off")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=300, bbox_inches="tight", pad_inches=0.05)
    plt.close(fig)


# Rows per catalog image. A handful of recurring architectures read fine
# packed together; past this, each extra row buys less than it costs in
# per-image legibility, so later rows spill onto numbered sibling pages.
_ROWS_PER_IMAGE = 6


def _plot_catalog_pages(
    catalog_rows: list[dict],
    representative_genes: dict[str, list[dict]],
    base_path: Path,
    display: dict,
    rows_per_image: int = _ROWS_PER_IMAGE,
    max_rows: int = 80,
) -> list[Path]:
    """Page a catalog at ``rows_per_image`` rows so no image is squeezed.

    A single page keeps ``base_path``'s name; more than one page gets it
    suffixed ``_page1``, ``_page2``, ...
    """
    rows = catalog_rows[:max_rows]
    if not rows:
        return []
    chunks = [rows[i:i + rows_per_image] for i in range(0, len(rows), rows_per_image)]
    written: list[Path] = []
    for page, chunk in enumerate(chunks, start=1):
        out_path = (
            base_path if len(chunks) == 1
            else base_path.with_name(f"{base_path.stem}_page{page}{base_path.suffix}")
        )
        _plot_catalog(chunk, representative_genes, out_path, display)
        written.append(out_path)
    return written


def _write_matrix(
    path: Path,
    summary_rows: list[dict],
    ordered_by_genome: dict[str, list[dict]],
    catalog_rows: list[dict],
    display: dict,
    locus_id: str = "",
) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    rank = {
        row["structure_id"]: index
        for index, row in enumerate(catalog_rows)
    }
    rows = sorted(
        summary_rows,
        key=lambda row: (
            rank.get(row["structure_id"], len(rank)),
            row["species"],
            row["strain"],
            row["genome_acc"],
        ),
    )
    fixed = list(MATRIX_FIXED_HEADERS)
    displayed = {
        row["genome_acc"]: _display_genes(
            ordered_by_genome.get(row["genome_acc"], [])
        )
        for row in rows
    }
    max_genes = max((len(genes) for genes in displayed.values()), default=0)
    headers = fixed + [
        f"g{index:03d}"
        for index in range(1, max_genes + 1)
    ]

    # Styling mirrors write_loci_xlsx() in writers.py; update both together.
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (locus_id or "cassette_matrix")[:31]
    sheet.append(headers)
    for summary in rows:
        genes = displayed[summary["genome_acc"]]
        values = [summary.get(column, "") for column in fixed]
        values.extend(
            f"[{_display_key(gene)}] {gene.get('product', '')}".strip()
            for gene in genes
        )
        values.extend([""] * (max_genes - len(genes)))
        sheet.append(values)
        excel_row = sheet.max_row
        status_value = str(summary.get("locus_status") or "").strip().upper()
        for index in range(1, len(fixed) + 1):
            cell = sheet.cell(excel_row, index)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", wrap_text=False)
            cell.font = Font(size=9)
            # Same status tint as loci.xlsx, so a row reads alike on either grid.
            if fixed[index - 1] == "locus_status" and status_value in STATUS_TINT:
                cell.fill = PatternFill("solid", fgColor=STATUS_TINT[status_value])
                cell.font = Font(size=9, bold=True)
        for index, gene in enumerate(genes, start=len(fixed) + 1):
            cell = sheet.cell(excel_row, index)
            style = _style_for_gene(gene, display)
            is_pseudo = str(gene.get("is_pseudo") or "").upper() in {
                "Y",
                "TRUE",
                "1",
            }
            color = "F4CCCC" if is_pseudo else _hex_color(style["color"])
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", wrap_text=False)
            cell.font = Font(italic=is_pseudo, size=9)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.comment = Comment(
                "\n".join([
                    f"locus_tag: {gene.get('locus_tag', '')}",
                    f"family: {gene.get('family', '')}",
                    f"zone: {gene.get('zone', '')}",
                    f"state: {gene.get('state', '')}",
                    f"piece: {gene.get('piece_idx', '')}",
                    f"product: {gene.get('product', '')}",
                    f"tblastn_pid: {gene.get('tblastn_pid', '')}",
                    f"tblastn_cov: {gene.get('tblastn_cov', '')}",
                ]),
                "LocusRuler",
            )
        for index in range(len(fixed) + len(genes) + 1, len(headers) + 1):
            sheet.cell(excel_row, index).border = thin

    header_fill = PatternFill("solid", fgColor="2F4F4F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.freeze_panes = freeze_after_status(fixed)
    widths = [MATRIX_FIXED_WIDTHS.get(name, 14) for name in fixed]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[
            sheet.cell(1, index).column_letter
        ].width = width
    for index in range(len(fixed) + 1, len(headers) + 1):
        sheet.column_dimensions[
            sheet.cell(1, index).column_letter
        ].width = 30
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _catalog(summary_rows: list[dict]) -> list[dict]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for row in summary_rows:
        if row["structure_id"]:
            groups[row["structure_id"]].append(row)
    rows = []
    for structure_id, members in groups.items():
        annotation_counts = Counter(row["annotation_signature"] for row in members)
        assembly_counts = Counter(row["assembly_signature"] for row in members)
        rows.append({
            "structure_id": structure_id,
            "structure_signature": members[0]["structure_signature"],
            "n_genomes": len(members),
            "n_annotation_variants": len(annotation_counts),
            "n_assembly_variants": len(assembly_counts),
            "most_common_annotation_signature": annotation_counts.most_common(1)[0][0],
            "most_common_assembly_signature": assembly_counts.most_common(1)[0][0],
            "example_genome": sorted(row["genome_acc"] for row in members)[0],
            "genera": ";".join(sorted({row["genus"] for row in members if row["genus"]})),
        })
    return sorted(rows, key=lambda row: (-int(row["n_genomes"]), row["structure_signature"]))


def run_cassette_structure(
    cfg_path: Path,
    settings_path: Path,
    target_name: str | None = None,
    output_root_override: Path | None = None,
    max_plot_rows: int = 80,
) -> dict[str, Path]:
    cfg_path = cfg_path.resolve()
    settings_path = settings_path.resolve()
    locus_cfg = load_locus_cfg(cfg_path)
    settings = load_settings(settings_path)
    locus_id = locus_cfg["locus_id"]
    target_name = resolve_target_name(settings, target_name, locus_cfg)
    target_cfg = get_target_cfg(settings, target_name)
    display = _display_settings(locus_cfg)
    output_root = (
        output_root_override
        if output_root_override is not None
        else Path(settings["paths"]["output_root"])
    )
    locus_out = output_root / target_name / locus_id
    out_dir = tables_dir(locus_out)

    reference_bp = reference_length(locus_cfg)
    cassette_span_bp = resolve_window(
        locus_cfg.get("cassette_max_span_bp", DERIVE),
        reference_bp, CASSETTE_SPAN_MULTIPLE,
    )
    print(f"[cassette] window - {describe_window('cassette_max_span', locus_cfg.get('cassette_max_span_bp', DERIVE), cassette_span_bp, reference_bp)}")

    genome_rows = [
        row for row in _read_csv(find_output(locus_out, "genome_summary.csv"))
        if row.get("locus_id") == locus_id
    ]
    diagnostic_rows = [
        row for row in _read_csv(find_output(locus_out, "gene_diagnostics.csv"))
        if row.get("locus_id") == locus_id
    ]
    strand_by_gene = _load_gene_strands(
        Path(target_cfg["db"]),
        diagnostic_rows,
    )
    for row in diagnostic_rows:
        row["strand"] = strand_by_gene.get(
            (row.get("genome_acc", ""), row.get("locus_tag", "")),
            row.get("strand", ""),
        )
    piece_rows = _read_csv(find_output(locus_out, "pieces.csv"))
    pieces_by_genome = _accepted_pieces(piece_rows, locus_id)
    diagnostics_by_genome: dict[str, list[dict]] = defaultdict(list)
    for row in diagnostic_rows:
        diagnostics_by_genome[row["genome_acc"]].append(row)

    # Bound the cassette by its flanks, falling back to accepted pieces.
    wanted: dict[str, tuple[str, int, int, int, int]] = {}
    for accession, rows in diagnostics_by_genome.items():
        pieces_here, _ = _colocated_pieces(
            pieces_by_genome.get(accession, {}), cassette_span_bp)
        span = extend_to_flanks(rows, pieces_here, reference_bp)
        if span is not None:
            wanted[accession] = span
    interval_by_genome = genes_in_span(Path(target_cfg["db"]), wanted)
    ref_rank = reference_rank(locus_cfg.get("_auto", {}).get("anchors", []))
    print("[cassette] reach - "
          + describe_intervals(wanted, len(genome_rows), reference_bp))

    ordered_by_genome: dict[str, list[dict]] = {}
    summary_rows: list[dict] = []
    all_families: set[str] = set()
    for genome in genome_rows:
        accession = genome["genome_acc"]
        pieces, offlocus_pieces = _colocated_pieces(
            pieces_by_genome.get(accession, {}), cassette_span_bp)
        ordered = _ordered_genes(diagnostics_by_genome.get(accession, []), pieces)

        span_genes = interval_by_genome.get(accession)
        if span_genes:
            # A union: the extension may only add.
            ordered = _renumber(augment_with_span(ordered, span_genes), ref_rank)
        ordered_by_genome[accession] = ordered
        members = [gene for gene in ordered if gene["is_cassette_member"] == "Y"]
        structure, annotation, state, assembly = _signatures(ordered)
        if not pieces:
            structure_status = "absent"
            structure = annotation = state = assembly = "absent"
        elif not members:
            structure_status = "empty"
        else:
            structure_status = "observed"
        sid = _structure_id(structure)
        families = sorted({_family_token(gene) for gene in members if _family_token(gene) != "unassigned"})
        all_families.update(families)
        contigs = [
            piece.get("contig", "")
            for piece in sorted(pieces.values(), key=_piece_sort_key)
            if piece.get("contig")
        ]
        summary_rows.append({
            "genome_acc": accession,
            "species": genome.get("species", ""),
            "strain": genome.get("strain", ""),
            "assembly_level": genome.get("assembly_level", ""),
            "genus": str(genome.get("species") or "").split()[0] if genome.get("species") else "",
            "locus_id": locus_id,
            "locus_status": genome.get("status", ""),
            "structure_status": structure_status,
            "structure_id": sid,
            "structure_signature": structure,
            "annotation_signature": annotation,
            "state_signature": state,
            "assembly_signature": assembly,
            "n_accepted_pieces": len(pieces),
            "n_offlocus_pieces": len(offlocus_pieces),
            "n_cassette_genes": len(members),
            "n_assigned_genes": sum(_family_token(gene) != "unassigned" for gene in members),
            "n_unassigned_genes": sum(_family_token(gene) == "unassigned" for gene in members),
            "families_present": ";".join(families),
            "accepted_contigs": "|".join(contigs),
            "membership_rule": "main_locus_ruler_accepted_pieces",
        })

    for row in summary_rows:
        present = set(filter(None, row["families_present"].split(";")))
        for family in sorted(all_families):
            row[f"has_{family}"] = "Y" if family in present else ""

    summary_by_genome = {row["genome_acc"]: row for row in summary_rows}
    gene_rows: list[dict] = []
    for accession, ordered in ordered_by_genome.items():
        sid = summary_by_genome[accession]["structure_id"]
        for gene in ordered:
            row = {
                "genome_acc": accession,
                "species": gene.get("species", summary_by_genome[accession]["species"]),
                "strain": gene.get("strain", summary_by_genome[accession]["strain"]),
                "locus_id": locus_id,
                "structure_id": sid,
                "structure_order": gene["structure_order"],
                "piece_order": gene["piece_order"],
                "within_piece_order": gene["within_piece_order"],
                "piece_idx": gene.get("piece_idx", ""),
                "piece_contig": gene.get("piece_contig", ""),
                "piece_q_start": gene.get("piece_q_start", ""),
                "piece_q_end": gene.get("piece_q_end", ""),
                "piece_strand": gene.get("piece_strand", ""),
                "zone": gene.get("zone", ""),
                "is_cassette_member": gene["is_cassette_member"],
                "membership_source": gene["membership_source"],
                "locus_tag": gene.get("locus_tag", ""),
                "gene_contig": gene.get("gene_contig", ""),
                "gene_start": gene.get("gene_start", ""),
                "gene_end": gene.get("gene_end", ""),
                "strand": gene.get("strand", ""),
                "family": gene.get("family", ""),
                "label": gene.get("label", ""),
                "status_role": gene.get("status_role", ""),
                "product": gene.get("product", ""),
                "state": gene.get("state", ""),
                "is_pseudo": gene.get("is_pseudo", ""),
                "tblastn_cov": gene.get("tblastn_cov", ""),
                "tblastn_pid": gene.get("tblastn_pid", ""),
            }
            gene_rows.append(row)

    catalog_rows = _catalog(summary_rows)
    representatives: dict[str, list[dict]] = {}
    for catalog in catalog_rows:
        candidates = [
            row for row in summary_rows if row["structure_id"] == catalog["structure_id"]
        ]
        representative = sorted(
            candidates,
            key=lambda row: (
                row["n_accepted_pieces"],
                row["n_unassigned_genes"],
                row["genome_acc"],
            ),
        )[0]
        catalog["representative_genome"] = representative["genome_acc"]
        catalog["representative_species"] = representative["species"]
        catalog["representative_strain"] = representative["strain"]
        representatives[catalog["structure_id"]] = ordered_by_genome[representative["genome_acc"]]

    # Step 6 tables go to diagnostics/; its figure sits beside the locus figure.
    summary_path = out_dir / "cassette_summary.csv"
    genes_path = out_dir / "cassette_genes.csv"
    catalog_path = out_dir / "structure_catalog.csv"
    figure_path = locus_out / "cassette_structure.png"
    singleton_figure_path = locus_out / "cassette_structure_singletons.png"
    matrix_path = out_dir / "cassette_matrix.xlsx"
    manifest_path = diagnostics_dir(locus_out) / "cassette_run_manifest.json"
    _write_csv(summary_path, summary_rows)
    _write_csv(genes_path, gene_rows)
    _write_csv(catalog_path, catalog_rows)
    # A structure seen in only one genome is a one-off, not a recurring
    # architecture -- keep it out of the main catalog plot so a single rare
    # rearrangement can't dominate the page, and page it separately instead.
    recurring_rows = [row for row in catalog_rows if row["n_genomes"] >= 2]
    singleton_rows = [row for row in catalog_rows if row["n_genomes"] == 1]
    figure_paths = _plot_catalog_pages(
        recurring_rows, representatives, figure_path, display,
        max_rows=max_plot_rows,
    )
    figure_paths += _plot_catalog_pages(
        singleton_rows, representatives, singleton_figure_path, display,
        max_rows=max_plot_rows,
    )
    _write_matrix(
        matrix_path,
        summary_rows,
        ordered_by_genome,
        catalog_rows,
        display,
        locus_id,
    )
    manifest_path.write_text(json.dumps({
        "schema_version": 2,
        "mode": "structure_discovery",
        "locus_id": locus_id,
        "target": target_name,
        "membership_rule": "main_locus_ruler_accepted_pieces",
        "biological_traits_used_for_membership": False,
        "whole_genome_off_contig_rescue": False,
        "outputs": {
            "summary": summary_path.name,
            "genes": genes_path.name,
            "catalog": catalog_path.name,
            "figures": [path.name for path in figure_paths],
            "matrix": matrix_path.name,
        },
    }, indent=2), encoding="utf-8")

    for obsolete in (
        "cassette_intervals.csv",
        "cargo_evidence.csv",
        "cassette_report.md",
        "classification_trace.csv",
        "cassette_structures.csv",
        "internal_genes.csv",
        "synteny_intervals.csv",
        "cassette_overview.csv",
        "OVERVIEW.md",
    ):
        path = out_dir / obsolete
        if path.exists():
            path.unlink()

    print(f"[cassette] membership: accepted LocusRuler pieces only")
    print(f"[cassette] genomes: {len(summary_rows)}")
    print(f"[cassette] structures: {len(catalog_rows)}")
    print(f"[cassette] summary -> {summary_path}")
    print(f"[cassette] genes   -> {genes_path}")
    print(f"[cassette] catalog -> {catalog_path}")
    for path in figure_paths:
        print(f"[cassette] figure  -> {path}")
    print(f"[cassette] matrix  -> {matrix_path}")
    return {
        "summary": summary_path,
        "genes": genes_path,
        "catalog": catalog_path,
        "figures": figure_paths,
        "matrix": matrix_path,
        "manifest": manifest_path,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Discover cassette structures from accepted LocusRuler pieces"
    )
    parser.add_argument("--config", required=True, help="Locus JSON config")
    parser.add_argument("--settings", default="settings.toml",
                        help="Path to settings.toml (default: settings.toml "
                             "in the current directory)")
    parser.add_argument("--target",
                        help="Default: the config's reference target, or "
                             "the only declared target; with several and no "
                             "target given, you are asked")
    parser.add_argument("--output-dir",
                        help="Override output_root from settings (default: "
                             "settings paths.output_root)")
    parser.add_argument("--max-plot-rows", type=int, default=80,
                        help="Cap on rows rendered per catalog (recurring, "
                             "singleton), before paging at 6 rows/image "
                             "(default: 80)")
    args = parser.parse_args()
    settings_path = Path(args.settings)
    if not settings_path.exists():
        raise SystemExit(
            f"ERROR: settings file not found: {settings_path}\n"
            f"       Pass --settings <path>, or run from the directory "
            f"holding settings.toml.")
    run_cassette_structure(
        Path(args.config),
        settings_path,
        target_name=args.target,
        output_root_override=Path(args.output_dir) if args.output_dir else None,
        max_plot_rows=args.max_plot_rows,
    )


if __name__ == "__main__":
    main()
