#!/usr/bin/env python3
"""Consolidated formatted workbook for one run."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from pathlib import Path

_PKG_DIR = Path(__file__).resolve().parent
if str(_PKG_DIR) not in sys.path:
    sys.path.insert(0, str(_PKG_DIR))

from writers import find_output

# Same palette as loci.xlsx so the two workbooks read as one system.
STATUS_COLOR = {
    "COMPLETE":    "E2EFDA",
    "CONDITIONAL": "DDEBF7",
    "DIVERGENT":   "EADCF8",
    "DECAYED":     "FFF2CC",
    "ABSENT":      "D9D9D9",
    "UNKNOWN":     "F3F3F3",
}
STATUS_ORDER = ["COMPLETE", "CONDITIONAL", "DIVERGENT", "DECAYED", "ABSENT", "UNKNOWN"]

HEADER_BG = "2F4F5F"
HEADER_FG = "FFFFFF"

MAX_COL_WIDTH = 46
MIN_COL_WIDTH = 8

# (sheet title, source filename, delimiter, description, column projection)
SHEETS = [
    ("Locus results", "genome_summary.csv", ",",
     "One row per genome: final status, coverage, gene counts, flank state.",
     [
         ("genome_acc", "Accession"),
         ("species", "Species"),
         ("strain", "Strain"),
         ("status", "Status"),
         ("coverage", "Coverage"),
         ("genes_found", "Genes found"),
         ("genes_expected", "Genes expected"),
         ("pseudo_count", "Pseudogenes"),
         ("cluster_bp", "Cluster bp"),
         ("synteny_class", "Synteny"),
         ("L_state", "L flank"),
         ("R_state", "R flank"),
         ("inversion", "Inverted"),
         ("fragmentation_type", "Fragmentation"),
         ("contig", "Contig"),
         ("notes", "Notes"),
     ]),
    # No separate cassette summary sheet; the Cassette order grid already carries those columns.
    ("Pieces", "pieces.csv", ",",
     "One row per accepted locus piece after duplication/paralog pruning.",
     [
         ("genome_acc", "Accession"),
         ("species", "Species"),
         ("strain", "Strain"),
         ("status", "Status"),
         ("piece_idx", "Piece"),
         ("contig", "Contig"),
         ("strand", "Strand"),
         ("s_lo", "Start"),
         ("s_hi", "End"),
         ("aligned_bp", "Aligned bp"),
         ("identity_pct", "Identity %"),
         ("n_covered_genes", "Genes"),
         ("covered_genes", "Covered genes"),
         ("internal_breaks", "Internal breaks"),
         ("contig_edge_proximal", "At contig edge"),
     ]),
    ("Breakpoints", "clade_markers.tsv", "\t",
     "Reference positions where many genomes break in the same place. Shared "
     "breakpoints tend to mark a clade.",
     None),
]


def _read_table(path: Path, delimiter: str) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8", errors="replace", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        try:
            header = next(reader)
        except StopIteration:
            return [], []
        return header, [row for row in reader]


def _coerce(value: str):
    """Turn a CSV string into a number when it unambiguously is one."""
    if value in ("", None):
        return None, None
    text = value.strip()
    if text.lower() in ("y", "yes", "true"):
        return text, None
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100.0, "0.0%"
        except ValueError:
            return value, None
    # Leave accessions, coordinates with commas, and version strings alone.
    if text.count(".") > 1 or "," in text:
        return value, None
    try:
        return int(text), None
    except ValueError:
        pass
    try:
        return float(text), None
    except ValueError:
        return value, None


def _autosize(worksheet, header, rows, get_column_letter):
    for index, name in enumerate(header, start=1):
        widest = len(str(name))
        for row in rows[:400]:          # sampling 400 rows is enough to size
            if index - 1 < len(row):
                widest = max(widest, len(str(row[index - 1])))
        width = min(MAX_COL_WIDTH, max(MIN_COL_WIDTH, widest + 2))
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_sheet(workbook, title, header, rows, description):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(title=title)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, size=10)

    for index, name in enumerate(header, start=1):
        cell = worksheet.cell(row=1, column=index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Works before and after a projection renames the column.
    status_index = next(
        (i + 1 for i, name in enumerate(header)
         if str(name).strip().lower() in ("status", "locus_status", "locus status")),
        None,
    )

    for row_number, row in enumerate(rows, start=2):
        status_value = ""
        if status_index and status_index - 1 < len(row):
            status_value = (row[status_index - 1] or "").strip().upper()
        fill = (
            PatternFill("solid", fgColor=STATUS_COLOR[status_value])
            if status_value in STATUS_COLOR else None
        )
        for index, value in enumerate(row, start=1):
            coerced, number_format = _coerce(value)
            cell = worksheet.cell(row=row_number, column=index, value=coerced)
            if number_format:
                cell.number_format = number_format
            if fill is not None and index == status_index:
                cell.fill = fill
                cell.font = Font(bold=True, size=10)

    worksheet.freeze_panes = "A2"
    if rows:
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
        )
    _autosize(worksheet, header, rows, get_column_letter)
    worksheet.sheet_properties.tabColor = HEADER_BG
    return worksheet, description


def _copy_formatted_sheet(workbook, source_path, title):
    """Copy an already-formatted worksheet into this workbook, styles intact."""
    from copy import copy

    from openpyxl import load_workbook
    from openpyxl.comments import Comment

    try:
        source_wb = load_workbook(source_path)
    except Exception as exc:
        print(f"[report] WARNING: could not read {source_path.name} ({exc}); "
              f"skipping the '{title}' sheet")
        return None

    try:
        source = source_wb.active
        target = workbook.create_sheet(title=title)

        for row in source.iter_rows():
            for cell in row:
                if cell.value is None and not (
                    cell.fill and cell.fill.fill_type == "solid"
                ):
                    continue
                new = target.cell(row=cell.row, column=cell.column, value=cell.value)
                new.font = copy(cell.font)
                new.fill = copy(cell.fill)
                new.border = copy(cell.border)
                new.alignment = copy(cell.alignment)
                new.number_format = cell.number_format
                if cell.comment is not None:
                    new.comment = Comment(cell.comment.text, cell.comment.author or "")

        for key, dimension in source.column_dimensions.items():
            if dimension.width:
                target.column_dimensions[key].width = dimension.width
        target.freeze_panes = source.freeze_panes or "A2"
        target.sheet_properties.tabColor = HEADER_BG
        return max(source.max_row - 1, 0)
    finally:
        source_wb.close()


STRUCTURE_COLUMNS = [
    ("structure_id", "Structure ID"),
    ("structure_signature", "Cassette structure"),
    ("n_genomes", "Genomes"),
    ("representative_species", "Example species"),
    ("representative_strain", "Example strain"),
]


def _read_structures(locus_dir):
    """Structure catalog rows for the Overview sheet."""
    source = find_output(locus_dir, "structure_catalog.csv")
    if not source.exists():
        return [], []
    header, rows = _read_table(source, ",")
    if not header:
        return [], []
    keep = [(header.index(src), label)
            for src, label in STRUCTURE_COLUMNS if src in header]
    if not keep:
        return [], []
    labels = [label for _, label in keep]
    projected = [
        [row[i] if i < len(row) else "" for i, _ in keep]
        for row in rows
    ]
    # Busiest structures first; the catalog is not sorted by genome count.
    count_at = next((n for n, (_, label) in enumerate(keep) if label == "Genomes"), None)
    if count_at is not None:
        def _count(row):
            try:
                return -int(row[count_at])
            except (TypeError, ValueError):
                return 0
        projected.sort(key=_count)
    return labels, projected


def _write_overview(workbook, locus_dir, locus_id, built, skipped, status_counts):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet = workbook.create_sheet(title="Overview", index=0)
    title_font = Font(bold=True, size=14)
    head_font = Font(bold=True, size=11, color=HEADER_FG)
    head_fill = PatternFill("solid", fgColor=HEADER_BG)

    worksheet["A1"] = "LocusRuler run report"
    worksheet["A1"].font = title_font
    worksheet["A2"] = f"locus_id: {locus_id}"
    worksheet["A3"] = f"source:   {locus_dir}"
    for ref in ("A2", "A3"):
        worksheet[ref].font = Font(italic=True, size=10)

    row = 5
    worksheet.cell(row=row, column=1, value="Final status").font = head_font
    worksheet.cell(row=row, column=1).fill = head_fill
    worksheet.cell(row=row, column=2, value="Genomes").font = head_font
    worksheet.cell(row=row, column=2).fill = head_fill
    worksheet.cell(row=row, column=3, value="Share").font = head_font
    worksheet.cell(row=row, column=3).fill = head_fill

    total = sum(status_counts.values()) or 1
    ordered = [s for s in STATUS_ORDER if s in status_counts]
    ordered += [s for s in sorted(status_counts) if s not in STATUS_ORDER]
    for status in ordered:
        row += 1
        count = status_counts[status]
        worksheet.cell(row=row, column=1, value=status).fill = PatternFill(
            "solid", fgColor=STATUS_COLOR.get(status, "FFFFFF")
        )
        worksheet.cell(row=row, column=1).font = Font(bold=True)
        worksheet.cell(row=row, column=2, value=count)
        share = worksheet.cell(row=row, column=3, value=count / total)
        share.number_format = "0.0%"
    row += 1
    worksheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    worksheet.cell(row=row, column=2, value=sum(status_counts.values())).font = Font(bold=True)

    structure_header, structure_rows = _read_structures(locus_dir)
    if structure_rows:
        row += 2
        worksheet.cell(row=row, column=1,
                       value="Cassette structures").font = head_font
        worksheet.cell(row=row, column=1).fill = head_fill
        worksheet.cell(
            row=row, column=5,
            value=f"{len(structure_rows)} distinct arrangements across "
                  f"{sum(status_counts.values())} genomes",
        ).font = Font(italic=True, size=10)
        row += 1
        for index, label in enumerate(structure_header, start=1):
            cell = worksheet.cell(row=row, column=index, value=label)
            cell.font = Font(bold=True, size=10)
        for entry in structure_rows:
            row += 1
            for index, value in enumerate(entry, start=1):
                coerced, number_format = _coerce(value)
                cell = worksheet.cell(row=row, column=index, value=coerced)
                if number_format:
                    cell.number_format = number_format

    row += 2
    worksheet.cell(row=row, column=1, value="Sheets in this workbook").font = head_font
    worksheet.cell(row=row, column=1).fill = head_fill
    worksheet.cell(row=row, column=2, value="Rows").font = head_font
    worksheet.cell(row=row, column=2).fill = head_fill
    worksheet.cell(row=row, column=5, value="Contents").font = head_font
    worksheet.cell(row=row, column=3).fill = head_fill
    for title, n_rows, description in built:
        row += 1
        worksheet.cell(row=row, column=1, value=title).font = Font(bold=True)
        worksheet.cell(row=row, column=2, value=n_rows)
        worksheet.cell(row=row, column=5, value=description)

    titles = {title for title, _, _ in built}
    if {"Locus genes", "Cassette genes"} <= titles:
        row += 2
        worksheet.cell(row=row, column=1,
                       value="The two gene grids").font = head_font
        worksheet.cell(row=row, column=1).fill = head_fill

        for label, note in (
            ("Locus genes",
             "Everything found in the locus neighborhood: the flanking genes "
             "that bracket it, the reference-family genes, anything unassigned "
             "in between, and a separator cell wherever the locus is split "
             "across pieces. Each cell also carries the evidence behind the "
             "call, e.g. [GeneB|protein] or [GeneA|protein|boundary]."),
            ("Cassette genes",
             "Only the genes judged to belong to the cassette. Flanks, "
             "unassigned neighbors and anything past a piece boundary are "
             "dropped, so a row is the arrangement itself. Typically about "
             "half as many cells as the same row on Locus genes."),
            ("Both",
             "A column is the Nth cell of THAT genome's own row, not a fixed "
             "reference position: a duplicated gene or a separator shifts "
             "everything after it. Read a row across; do not compare a column "
             "down the sheet. Use 'Cassette genes' structure ids, or the "
             "catalog above, to group genomes by arrangement."),
        ):
            row += 1
            cell = worksheet.cell(row=row, column=1, value=label)
            cell.font = Font(bold=True)
            worksheet.cell(row=row, column=5, value=note)

    if skipped:
        row += 2
        worksheet.cell(row=row, column=1, value="Not present in this run").font = head_font
        worksheet.cell(row=row, column=1).fill = head_fill
        for name in skipped:
            row += 1
            worksheet.cell(row=row, column=1, value=name)
            worksheet.cell(row=row, column=5, value="source file absent")

    row += 2
    worksheet.cell(row=row, column=1, value="Left outside this workbook").font = head_font
    worksheet.cell(row=row, column=1).fill = head_fill
    for name, why in (
        ("diagnostics/gene_diagnostics.csv",
         "Every displayed gene cell with the evidence behind its label. Carries "
         "full nt_seq/aa_seq, where a long gene can approach Excel's "
         "32,767-character limit."),
        ("diagnostics/hsp_diagnostics.csv",
         "HSP-level audit trail, tens of thousands of rows per run."),
        ("tables/cassette_summary.csv",
         "The wide form behind 'Cassette order': per-family has_* columns, the "
         "annotation/state/assembly signatures, and piece counts."),
        ("tables/cassette_genes.csv",
         "Gene-level backing for 'Cassette order'."),
        ("tables/marker_matrix.csv",
         "The Breakpoints sheet as a binary genome-by-position matrix, shaped "
         "for phylogeny tools rather than for reading."),
        ("diagnostics/domain_recovery_diagnostics.csv",
         "Genes rescued by the optional Pfam/HMM layer; empty unless "
         "[domain_recovery] is enabled."),
        ("cluster_heatmap.png / cassette_structure*.png",
         "The figures, at the top level. Cassette structure is split into "
         "recurring-vs-singleton catalogs, each paged at 6 rows/image."),
    ):
        row += 1
        worksheet.cell(row=row, column=1, value=name)
        worksheet.cell(row=row, column=5, value=why)

    for column, width in (("A", 34), ("B", 46), ("C", 11),
                          ("D", 26), ("E", 88)):
        worksheet.column_dimensions[column].width = width
    for cells in worksheet.iter_rows(min_col=5, max_col=5):
        for cell in cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.sheet_properties.tabColor = "C00000"
    return worksheet


def build_report(locus_dir: Path, out_path: Path, locus_id: str | None = None) -> Path:
    """Write a consolidated formatted workbook for one locus run."""
    from openpyxl import Workbook

    locus_dir = Path(locus_dir)
    out_path = Path(out_path)
    if not locus_dir.is_dir():
        raise FileNotFoundError(f"locus dir not found: {locus_dir}")
    locus_id = locus_id or locus_dir.name

    workbook = Workbook()
    workbook.remove(workbook.active)

    built: list[tuple[str, int, str]] = []
    skipped: list[str] = []
    status_counts: Counter = Counter()

    for title, filename, delimiter, description, projection in SHEETS:
        source = find_output(locus_dir, filename)
        if not source.exists():
            skipped.append(filename)
            continue
        header, rows = _read_table(source, delimiter)
        if not header:
            skipped.append(filename)
            continue

        if filename == "genome_summary.csv" and "status" in header:
            index = header.index("status")
            status_counts.update(
                (row[index] or "UNKNOWN").strip().upper()
                for row in rows if index < len(row)
            )

        if projection:
            keep = [(header.index(src), label)
                    for src, label in projection if src in header]
            if keep:
                header = [label for _, label in keep]
                rows = [
                    [row[i] if i < len(row) else "" for i, _ in keep]
                    for row in rows
                ]

        _write_sheet(workbook, title, header, rows, description)
        built.append((title, len(rows), description))

    # Both matrices arrive already formatted, so copy them rather than rebuild them.
    for filename, title, note in (
        ("loci.xlsx", "Locus genes",
         "Every gene found at the locus, including flanks, unassigned "
         "neighbors and piece separators; each cell carries its evidence."),
        ("cassette_matrix.xlsx", "Cassette genes",
         "Only the genes judged to belong to the cassette, in order. Read a "
         "row across: a column is not a fixed reference position."),
    ):
        source = find_output(locus_dir, filename)
        if not source.exists():
            skipped.append(filename)
            continue
        n_rows = _copy_formatted_sheet(workbook, source, title)
        if n_rows is not None:
            built.append((title, n_rows, note))

    # Sheet order: the two matrices, then the per-genome answer, the wide metrics, then detail.
    desired = ["Locus genes", "Cassette genes", "Locus results",
               "Pieces", "Breakpoints"]
    for position, title in enumerate(t for t in desired if t in workbook.sheetnames):
        workbook.move_sheet(
            title, offset=position - workbook.sheetnames.index(title)
        )

    if not built:
        raise FileNotFoundError(
            f"no LocusRuler tables found under {locus_dir}; run the pipeline first"
        )

    # List the sheets on Overview in the order the tabs appear.
    tab_order = {name: i for i, name in enumerate(workbook.sheetnames)}
    built.sort(key=lambda entry: tab_order.get(entry[0], len(tab_order)))

    _write_overview(workbook, locus_dir, locus_id, built, skipped, status_counts)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    print(f"[report] locus_report.xlsx - {out_path}  "
          f"({len(built)} sheets, {sum(n for _, n, _ in built):,} data rows)")
    return out_path


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Build a consolidated formatted workbook for one LocusRuler run.",
    )
    parser.add_argument("--locus-dir", required=True,
                        help="output/<target>/<locus_id>")
    parser.add_argument("--out", default=None,
                        help="Output .xlsx path (default: <locus-dir>/locus_report.xlsx)")
    parser.add_argument("--locus-id", default=None,
                        help="Label for the Overview sheet (default: directory name)")
    args = parser.parse_args(argv)

    locus_dir = Path(args.locus_dir)
    out_path = Path(args.out) if args.out else locus_dir / "locus_report.xlsx"
    build_report(locus_dir, out_path, args.locus_id)
    return 0


if __name__ == "__main__":
    sys.exit(main())
