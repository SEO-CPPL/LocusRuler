#!/usr/bin/env python3
"""Output file writers."""

import csv
import sys
from pathlib import Path
from typing import Optional


# ── Sequence retrieval helpers ─────────────────────────────────────────────
_NT_COMPLEMENT = str.maketrans("ACGTacgt", "TGCAtgca")


def _revcomp(seq: str) -> str:
    return seq.translate(_NT_COMPLEMENT)[::-1]


def _find_fna_path(fna_dir: str, acc: str) -> Optional[Path]:
    """Return the genome FASTA path for *acc* in *fna_dir*, or None."""
    d = Path(fna_dir)
    for ext in (".fna.gz", ".fna", ".fa.gz", ".fa", ".fasta.gz", ".fasta"):
        # exact match first (e.g. GCF_000000000.1.fna.gz)
        p = d / f"{acc}{ext}"
        if p.exists():
            return p
        # NCBI-style: GCF_000000000.1_ASM000000v1_genomic.fna.gz
        candidates = list(d.glob(f"{acc}_*{ext}"))
        if candidates:
            return candidates[0]
    return None


def _load_genome_contigs(fna_dir: str, acc: str,
                         _cache: dict[str, Optional[dict[str, str]]] = {}) -> Optional[dict[str, str]]:
    """Cache-backed loader: returns {contig_id: sequence} or None."""
    if acc in _cache:
        return _cache[acc]
    fna_path = _find_fna_path(fna_dir, acc)
    if fna_path is None:
        _cache[acc] = None
        return None
    try:
        # Imported here to avoid a circular dependency.
        _pkg = Path(__file__).parent
        if str(_pkg) not in sys.path:
            sys.path.insert(0, str(_pkg))
        from db_utils import load_fna  # type: ignore
        _cache[acc] = load_fna(fna_path)
    except Exception as e:
        print(f"[writers] WARN: could not load FNA for {acc}: {e}")
        _cache[acc] = None
    return _cache[acc]


def _extract_nt_seq(contigs: dict[str, str],
                    contig: str, start, end, strand: str) -> str:
    """Extract coding nucleotide sequence from *contigs* (1-based GFF coords)."""
    seq = contigs.get(contig, "")
    if not seq or not start or not end:
        return ""
    try:
        s, e = int(start) - 1, int(end)   # convert to 0-based half-open
    except (TypeError, ValueError):
        return ""
    subseq = seq[s:e]
    if strand == "-":
        subseq = _revcomp(subseq)
    return subseq


# ── Output layout ──────────────────────────────────────────────────────────
TABLES_DIRNAME = "tables"
DIAGNOSTICS_DIRNAME = "diagnostics"

TABLE_FILES = (
    "genome_summary.csv",
    "loci.csv",
    "loci.xlsx",
    "pieces.csv",
    "clade_markers.tsv",
    "marker_matrix.csv",
    "cassette_summary.csv",
    "cassette_genes.csv",
    "cassette_matrix.xlsx",
    "structure_catalog.csv",
)

DIAGNOSTIC_FILES = (
    "gene_diagnostics.csv",
    "hsp_diagnostics.csv",
    "domain_recovery_diagnostics.csv",
    "cassette_run_manifest.json",
)

# Searched in order, so a file found in tables/ wins over a stale flat copy.
_SEARCH_DIRS = (TABLES_DIRNAME, DIAGNOSTICS_DIRNAME, "")


def tables_dir(output_dir) -> Path:
    """Return (and create) the results subdirectory for a locus output dir."""
    path = Path(output_dir) / TABLES_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def diagnostics_dir(output_dir) -> Path:
    """Return (and create) the audit-trail subdirectory for a locus output dir."""
    path = Path(output_dir) / DIAGNOSTICS_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def output_home(filename) -> str:
    """Which subdirectory a given output file belongs in."""
    if filename in DIAGNOSTIC_FILES:
        return DIAGNOSTICS_DIRNAME
    return TABLES_DIRNAME


def find_output(output_dir, filename):
    """Resolve an output file, searching tables/ then diagnostics/ then the root."""
    output_dir = Path(output_dir)
    for subdir in _SEARCH_DIRS:
        candidate = (output_dir / subdir / filename) if subdir else (output_dir / filename)
        if candidate.exists():
            return candidate
    return output_dir / output_home(filename) / filename


# Status tints, shared with the cassette matrix so a row reads the same way on either grid.
STATUS_TINT = {
    "COMPLETE":    "E2EFDA",
    "CONDITIONAL": "DDEBF7",
    "DIVERGENT":   "EADCF8",
    "DECAYED":     "FFF2CC",
    "ABSENT":      "D9D9D9",
    "UNKNOWN":     "D9D9D9",
}


# ── Shared grid layout (loci.xlsx and cassette_matrix.xlsx) ──────────────────────────────────────────────────────
IDENTITY_HEADERS = ["genome_acc", "species", "strain", "assembly_level"]
IDENTITY_WIDTHS = {"genome_acc": 22, "species": 30, "strain": 14,
                   "assembly_level": 14}
STATUS_ALIASES = ("status", "locus_status")

# loci.xlsx: identity, the call, then the counts behind it.
LOCI_FIXED_HEADERS = IDENTITY_HEADERS + [
    "status", "coverage", "cluster_bp",
    "genes_found", "pseudo_count", "contig",
]
LOCI_FIXED_WIDTHS = {**IDENTITY_WIDTHS, "status": 16, "coverage": 8,
                     "cluster_bp": 10, "genes_found": 10,
                     "pseudo_count": 10, "contig": 22}


def freeze_after_status(headers: list[str]) -> str:
    """Freeze reference pinning the identity columns and the status call."""
    from openpyxl.utils import get_column_letter

    index = next(i for i, name in enumerate(headers) if name in STATUS_ALIASES)
    return f"{get_column_letter(index + 2)}2"


# ── Cell color palette (loci.xlsx) ──────────────────────────────────────────────────────
_COL_CLUSTER   = "F4B942"   # orange      - intact reference cluster genes
_COL_FLANK     = "C9DAF8"   # light blue  - flank context (housekeeping)
_COL_SEPARATOR = "2F2F2F"   # near-black  - between-piece separator cell
_COL_PSEUDO    = "FF0000"   # red         - GFF-confirmed pseudogene
_COL_DNA_ONLY  = "FFF59D"   # pale yellow - DNA-only signal (no GFF gene)
_COL_DIVERGENT = "FFF59D"   # pale yellow - divergent gene signal
_COL_UNKNOWN   = "D9D9D9"   # gray


def zone_color(zone: Optional[str], is_pseudo: bool,
                divergent: bool = False, dna_only: bool = False) -> str:
    """Background hex for a gene cell:
       pseudo (GFF-confirmed) > dna_only > divergent > zone tint.
    """
    if zone == "separator":  return _COL_SEPARATOR
    if is_pseudo:            return _COL_PSEUDO
    if dna_only:             return _COL_DNA_ONLY
    if divergent:          return _COL_DIVERGENT
    if zone == "cluster":    return _COL_CLUSTER
    if zone in ("flank_left", "flank_right"): return _COL_FLANK
    return _COL_UNKNOWN


def _cell_label_parts(g: dict, res: dict) -> tuple[str, bool, bool, bool, str]:
    """Return (label, is_divergent, is_dna_style, strong, detail)."""
    fam      = g.get("family")
    is_ps    = bool(g.get("is_pseudo", False))
    dna_only = bool(g.get("is_dna_only", False))
    zone     = g.get("_zone")
    states_by_family  = res.get("_gene_states_by_family") or {}
    signals_by_family = res.get("_gene_signals_by_family") or {}
    relations_by_family = res.get("_gene_relations_by_family") or {}
    divergent_reasons   = res.get("_gene_divergent_reasons_by_family") or {}
    divergent_details   = res.get("_gene_divergent_details_by_family") or {}

    is_divergent = (zone == "cluster" and not is_ps and not dna_only and
                    states_by_family.get(fam) == "DIVERGENT")
    signal_word = ""
    if zone == "cluster" and not dna_only:
        sig = signals_by_family.get(fam, {})
        has_T = bool(sig.get("T"))
        has_D = bool(sig.get("D"))
        if has_D and not has_T:
            signal_word = "dna"
        elif has_T and not has_D:
            signal_word = "protein"

    parts: list[str] = []
    if dna_only:
        parts.append("dna")
    elif signal_word:
        parts.append(signal_word)
    if is_ps:
        parts.append("pseudo")
    elif is_divergent:
        parts.append(f"div:{divergent_reasons.get(fam, 'ambig')}")
    rel = relations_by_family.get(fam, {})
    piece_relation = g.get("_piece_relation") or rel.get("piece_relation")
    if piece_relation == "boundary_extension":
        parts.append("boundary")
    elif piece_relation == "adjacent_rescue":
        parts.append("adjacent")

    suffix = ("|" + "|".join(parts)) if parts else ""
    label = f"[{fam or 'unassigned'}{suffix}]"
    detail = divergent_details.get(fam, "") if is_divergent else ""
    strong = is_ps or dna_only or (signal_word == "dna")
    return label, is_divergent, (dna_only or signal_word == "dna"), strong, detail


def _loci_row_order(
    all_locus_genes: dict[str, list[dict]],
    genome_meta: dict[str, dict],
) -> list[str]:
    """Accessions grouped by which genes they carry, loci.csv and loci.xlsx alike."""
    def signature(acc: str) -> str:
        return ">".join(g.get("family") or "unassigned"
                        for g in all_locus_genes[acc] if g.get("_zone") == "cluster")

    def key(acc: str) -> tuple:
        return (signature(acc), genome_meta.get(acc, {}).get("species", ""), acc)

    return sorted(all_locus_genes, key=key)


# ── loci.csv ──────────────────────────────────────────────────────
def write_loci_csv(
    all_locus_genes: dict[str, list[dict]],
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
) -> None:
    locus_id   = locus_cfg["locus_id"]
    auto       = locus_cfg.get("_auto", {})
    n_expected = auto.get("n_genes_ref", 0)
    max_genes  = max((len(gs) for gs in all_locus_genes.values()), default=0)
    max_genes  = max(max_genes, n_expected)
    gene_cols  = [f"g{i+1:03d}" for i in range(max_genes)]

    header = (["genome_acc", "species", "strain", "assembly_level",
               "locus_id", "status", "coverage",
               "L2_L1_bp", "flank_bp", "R1_R2_bp",
               "cluster_bp", "order_flags",
               "genes_found", "pseudo_count", "contig"]
              + gene_cols)

    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        for acc in _loci_row_order(all_locus_genes, genome_meta):
            res   = ruler_results.get(acc, {})
            meta  = genome_meta.get(acc, {})
            genes = all_locus_genes[acc]
            gene_cells = []
            for g in genes:
                if g.get("_zone") == "separator":
                    gene_cells.append(g.get("product", "|||"))
                    continue
                label, *_ = _cell_label_parts(g, res)
                gene_cells.append(f"{label} {g['product'][:50]}")
            gene_cells += [""] * (max_genes - len(gene_cells))
            cov   = res.get("coverage")
            cov_s = f"{cov:.1%}" if cov is not None else ""
            flags = res.get("order_flags") or []
            w.writerow([
                acc,
                meta.get("species", ""),
                meta.get("strain", ""),
                meta.get("assembly_level", ""),
                locus_id,
                res.get("status", ""),
                cov_s,
                res.get("L2_L1_bp") if res.get("L2_L1_bp") is not None else "",
                res.get("flank_bp", ""),
                res.get("R1_R2_bp") if res.get("R1_R2_bp") is not None else "",
                res.get("cluster_bp", ""),
                "; ".join(flags) if flags else "",
                res.get("genes_found", ""),
                res.get("pseudo_count", ""),
                res.get("contig", ""),
            ] + gene_cells)
    print(f"[content] loci.csv - {out_path}")


def write_gene_diagnostics_csv(
    all_locus_genes: dict[str, list[dict]],
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
    tgt_faa_path: str = "",
    tgt_faa_index: Optional[dict[str, int]] = None,
    fna_dir: str = "",
) -> None:
    """Machine-readable per-cell explanations for labels in loci.xlsx."""
    # Lazy-build FAA index if caller did not pass one.
    _faa_index: dict[str, int] = {}
    if tgt_faa_path and Path(tgt_faa_path).exists():
        if tgt_faa_index is not None:
            _faa_index = tgt_faa_index
        else:
            try:
                _pkg = Path(__file__).parent
                if str(_pkg) not in sys.path:
                    sys.path.insert(0, str(_pkg))
                from flank_blast import _make_faa_index  # type: ignore
                _faa_index = _make_faa_index(tgt_faa_path)
            except Exception as e:
                print(f"[writers] WARN: could not index FAA: {e}")

    _get_aa: Optional[object] = None
    if _faa_index:
        try:
            from flank_blast import _get_seq_from_faa  # type: ignore
            _get_aa = _get_seq_from_faa
        except Exception:
            pass

    headers = [
        "genome_acc", "species", "strain", "locus_id", "status",
        "gene_col", "zone", "piece_idx", "piece_contig", "piece_s_lo", "piece_s_hi",
        "locus_tag", "gene_contig", "gene_start", "gene_end",
        "family", "status_role", "label", "nt_seq", "aa_seq", "product", "state",
        "tblastn_cov", "tblastn_pid", "tblastn_hit_contig",
        "tblastn_hit_lo", "tblastn_hit_hi",
        "cluster_blastn_cov", "T", "D", "divergent_reason",
        "divergent_detail", "piece_relation", "outside_piece_gap_bp",
        "reference_extension_bp",
        "rescue_reason", "is_pseudo", "is_dna_only",
    ]
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        status_role_by_family = {
            (a.get("family") or a.get("locus_tag", "")): a.get("status_role", "")
            for a in locus_cfg.get("_auto", {}).get("anchors", [])
        }
        for acc in sorted(all_locus_genes):
            res = ruler_results.get(acc, {})
            meta = genome_meta.get(acc, {})
            states_by_family = res.get("_gene_states_by_family") or {}
            signals_by_family = res.get("_gene_signals_by_family") or {}
            metrics_by_family = res.get("_gene_metrics_by_family") or {}
            ranges_by_family = res.get("_gene_s_ranges_by_family") or {}
            relations_by_family = res.get("_gene_relations_by_family") or {}
            divergent_reasons = res.get("_gene_divergent_reasons_by_family") or {}
            divergent_details = res.get("_gene_divergent_details_by_family") or {}
            _contigs: Optional[dict[str, str]] = (
                _load_genome_contigs(fna_dir, acc) if fna_dir else None
            )
            for idx, g in enumerate(all_locus_genes[acc], start=1):
                if g.get("_zone") == "separator":
                    continue
                fam = g.get("family") or ""
                label, *_ = _cell_label_parts(g, res)
                sig = signals_by_family.get(fam, {})
                met = metrics_by_family.get(fam, {})
                rng = ranges_by_family.get(fam, {})
                rel = relations_by_family.get(fam, {})
                piece_relation = g.get("_piece_relation") or rel.get("piece_relation", "")
                outside_gap = g.get("_outside_piece_gap_bp")
                if outside_gap in (None, ""):
                    outside_gap = rel.get("outside_piece_gap_bp", "")
                reference_extension = g.get("_reference_extension_bp")
                if reference_extension in (None, ""):
                    reference_extension = rel.get("reference_extension_bp", "")
                rescue_reason = g.get("_rescue_reason") or rel.get("rescue_reason", "")
                lt = g.get("locus_tag", "")
                nt_seq = (
                    _extract_nt_seq(_contigs, g.get("contig", ""),
                                    g.get("start"), g.get("end"),
                                    g.get("strand", "+"))
                    if _contigs else ""
                )
                aa_seq = (
                    _get_aa(tgt_faa_path, lt, _faa_index)  # type: ignore[call-arg]
                    if (_get_aa and lt and _faa_index)
                    else ""
                ) or ""
                # Aux probe hits carry their own state/cov/pid on the gene dict.
                _has_aux_metrics = any(
                    k in g for k in ("_state", "_tblastn_cov", "_tblastn_pid")
                )
                _state = (g.get("_state") if _has_aux_metrics
                          else states_by_family.get(fam, ""))
                _cov = (g.get("_tblastn_cov") if _has_aux_metrics
                        else met.get("cov", ""))
                _pid = (g.get("_tblastn_pid") if _has_aux_metrics
                        else met.get("pid", ""))
                _hit_contig = (g.get("_piece_contig") if _has_aux_metrics
                               else rng.get("contig", ""))
                _hit_lo = (g.get("_piece_s_lo") if _has_aux_metrics
                           else rng.get("s_lo", ""))
                _hit_hi = (g.get("_piece_s_hi") if _has_aux_metrics
                           else rng.get("s_hi", ""))
                w.writerow([
                    acc,
                    meta.get("species", ""),
                    meta.get("strain", ""),
                    locus_cfg["locus_id"],
                    res.get("status", ""),
                    f"g{idx:03d}",
                    g.get("_zone", ""),
                    g.get("_piece_idx", ""),
                    g.get("_piece_contig", ""),
                    g.get("_piece_s_lo", ""),
                    g.get("_piece_s_hi", ""),
                    lt,
                    g.get("contig", ""),
                    g.get("start", ""),
                    g.get("end", ""),
                    fam or "unassigned",
                    status_role_by_family.get(fam, ""),
                    label,
                    nt_seq,
                    aa_seq,
                    g.get("product", ""),
                    _state,
                    _cov,
                    _pid,
                    _hit_contig,
                    _hit_lo,
                    _hit_hi,
                    met.get("cluster_dna_cov", ""),
                    "Y" if sig.get("T") else "",
                    "Y" if sig.get("D") else "",
                    divergent_reasons.get(fam, ""),
                    divergent_details.get(fam, ""),
                    piece_relation,
                    outside_gap,
                    reference_extension,
                    rescue_reason,
                    "Y" if g.get("is_pseudo") else "",
                    "Y" if g.get("is_dna_only") else "",
                ])
    print(f"[content] gene_diagnostics.csv - {out_path}")


# Backward-compatible alias for older callers/tests.
write_loci_diagnostics_csv = write_gene_diagnostics_csv


def write_hsp_diagnostics_csv(
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
) -> None:
    """Machine-readable HSP audit trail for piece assembly."""
    headers = [
        "genome_acc", "species", "strain", "locus_id", "status",
        "hsp_idx", "used_for_assembly", "exclude_reason",
        "weak_hsp_rescue", "rescue_reason", "rescue_gap_bp", "rescue_domain_hits",
        "piece_idx", "qseqid", "sseqid", "qstart", "qend", "sstart",
        "send", "pident", "length", "bitscore", "covered_ref_genes",
        "max_gene_cov", "overlap_details",
    ]

    def _fmt_details(details: list[dict]) -> tuple[str, str, str]:
        if not details:
            return "", "", ""
        parts = []
        max_cov = 0.0
        covered = []
        for d in details:
            cov = float(d.get("coverage") or 0.0)
            max_cov = max(max_cov, cov)
            lt = d.get("locus_tag", "")
            parts.append(f"{lt}:{cov:.3f}")
            covered.append(lt)
        return ";".join(covered), f"{max_cov:.3f}", ";".join(parts)

    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for acc in sorted(ruler_results):
            res = ruler_results.get(acc, {})
            meta = genome_meta.get(acc, {})
            hsp_to_piece: dict[int, int] = {}
            for p_idx, piece in enumerate(res.get("_pieces") or []):
                for h_idx in piece.get("_hsp_indices") or []:
                    try:
                        hsp_to_piece[int(h_idx)] = p_idx
                    except (TypeError, ValueError):
                        continue

            rows = []
            seen_hsp: set[int] = set()
            for h in res.get("_assembly_hsps") or []:
                if "_hsp_idx" in h:
                    rows.append(h)
                    seen_hsp.add(int(h.get("_hsp_idx", -1)))
            for h in res.get("_orphans") or []:
                if "_hsp_idx" in h and int(h.get("_hsp_idx", -1)) not in seen_hsp:
                    rows.append(h)
                    seen_hsp.add(int(h.get("_hsp_idx", -1)))
            for h in res.get("_assembly_excluded_hsps") or []:
                if "_hsp_idx" in h and int(h.get("_hsp_idx", -1)) not in seen_hsp:
                    rows.append(h)

            rows.sort(key=lambda h: int(h.get("_hsp_idx", 10**9)))
            for h in rows:
                h_idx = int(h.get("_hsp_idx", -1))
                covered, max_cov, details = _fmt_details(
                    h.get("_gene_overlap_details") or [])
                used = bool(h.get("_used_for_assembly"))
                w.writerow([
                    acc,
                    meta.get("species", ""),
                    meta.get("strain", ""),
                    locus_cfg["locus_id"],
                    res.get("status", ""),
                    h_idx if h_idx >= 0 else "",
                    "Y" if used else "",
                    h.get("_assembly_excluded_reason", ""),
                    "Y" if h.get("_weak_hsp_rescue") else "",
                    h.get("_rescue_reason", ""),
                    h.get("_rescue_gap_bp", ""),
                    ";".join(
                        f"{d.get('locus_tag','')}:{float(d.get('coverage') or 0):.3f}"
                        for d in (h.get("_rescue_domain_hits") or [])
                    ),
                    hsp_to_piece.get(h_idx, ""),
                    h.get("qseqid", ""),
                    h.get("sseqid", ""),
                    h.get("qstart", ""),
                    h.get("qend", ""),
                    h.get("sstart", ""),
                    h.get("send", ""),
                    h.get("pident", ""),
                    h.get("length", ""),
                    h.get("bitscore", ""),
                    covered,
                    max_cov,
                    details,
                ])
    print(f"[content] hsp_diagnostics.csv - {out_path}")


write_piece_hsp_diagnostics_csv = write_hsp_diagnostics_csv


def write_output_guide(out_path: Path) -> None:
    out_path.write_text(
        "\n".join([
            "# LocusRuler output guide",
            "",
            "## Layout",
            "```",
            "<locus_id>/",
            "  locus_report.xlsx        start here (locus-ruler-report, or run with --report)",
            "  cluster_heatmap.png      locus figure (locus-ruler-heatmap, or --heatmap)",
            "  cassette_structure.png   cassette figure (recurring structures; singletons",
            "                           split into cassette_structure_singletons.png; both",
            "                           page at 6 rows/image as _page1, _page2, ...)",
            "  OUTPUTS.md               this file",
            "  tables/                  results: what the workbook is built from",
            "  diagnostics/             the audit trail: why a call came out that way",
            "```",
            "",
            "## Open these",
            "- `locus_report.xlsx`: the whole run in one workbook. Sheets, in order: Overview (status counts + the cassette structure catalog), Locus genes, Cassette genes, Locus results, Pieces, Breakpoints.",
            "  - `Locus genes`: every gene found at the locus -- flanks, reference families, unassigned neighbors, and a separator wherever the locus is split across pieces. Cells carry the evidence, e.g. `[GeneB|protein]`.",
            "  - `Cassette genes`: only the genes judged to belong to the cassette, so a row is the arrangement itself. Usually about half the cells of the same row on `Locus genes`.",
            "  - On both, a column is the Nth cell of that genome's own row, not a fixed reference position: a duplicated gene or a separator shifts everything after it. Read rows across, not columns down. Group genomes by the structure id instead.",
            "- `cluster_heatmap.png` / `cassette_structure*.png`: the figures. Cassette structure is split into recurring-vs-singleton catalogs, each paged at 6 rows/image.",
            "",
            "## tables/ - the results",
            "Everything downstream scripts should read, and everything the workbook is built from. Step 6's cassette tables live here too: step 6 re-reads this run rather than computing a new one.",
            "",
            "- `genome_summary.csv`: one row per genome. The main result as plain CSV.",
            "- `loci.csv` / `loci.xlsx`: strain-by-reference-gene matrix, with compact labels such as `[fam]`, `[fam|protein]`, `[fam|dna]`, `[fam|div:pid]`, `[fam|protein|div:pid|boundary]`.",
            "- `pieces.csv`: accepted piece coordinates and covered reference genes after duplication/paralog pruning.",
            "- `clade_markers.tsv` and `marker_matrix.csv`: reference positions where many genomes break in the same place, and the same information as a binary genome-by-position matrix for phylogeny tools.",
            "- `cassette_summary.csv`, `cassette_genes.csv`, `cassette_matrix.xlsx`, `structure_catalog.csv`: step 6 output, per genome, per cassette gene, as a matrix, and per distinct structure.",
            "",
            "## diagnostics/ - why LocusRuler called it that",
            "Nothing downstream depends on these. They exist for when a result looks wrong.",
            "",
            "- `gene_diagnostics.csv`: one row per displayed gene cell with piece coordinates, label cause, tblastn metrics, cluster-blastn DNA coverage, T/D signals, divergent-state reasons, piece relation, rescue reason, and pseudo flags.",
            "- `hsp_diagnostics.csv`: HSP-level audit trail for piece assembly, including intergenic HSPs excluded from boundaries.",
            "- `domain_recovery_diagnostics.csv`: genes rescued by the optional Pfam/HMM layer (only when `[domain_recovery]` is enabled).",
            "- `cassette_run_manifest.json`: step 6 provenance.",
            "",
            "",
            "## Final status",
            "- `COMPLETE`: every CORE status cell is INTACT, and any SHARED cell is INTACT.",
            "- `CONDITIONAL`: every CORE status cell is INTACT, but at least one SHARED cell is not INTACT.",
            "- `DIVERGENT`: all CORE cells are present and at least one CORE cell is DIVERGENT.",
            "- `DECAYED`: the core cluster is not wholly absent, at least one CORE cell is PSEUDOGENE or ABSENT, and locus coverage passes the DECAYED floor.",
            "- `ABSENT`: every CORE cell is PSEUDOGENE or ABSENT.",
            "- Piece coordinates, coverage, piece count, and fragmentation fields remain diagnostic outputs.",
            "",
            "## Label notes",
            "- `[fam]`: normal cluster gene signal.",
            "- `[fam|protein]`: protein/tblastn signal without enough cluster-DNA signal in the accepted piece.",
            "- `[fam|dna]`: DNA-only signal.",
            "- `[fam|pseudo]`: GFF-confirmed pseudogene.",
            "- `[fam|div:pid]`: divergent because identity is below the intact threshold.",
            "- `[fam|div:cov]`: divergent because coverage is below the intact threshold.",
            "- `[fam|div:frag]`: divergent short protein fragment with DNA support.",
            "- `[fam|...|boundary]`: strong inner-anchor tblastn hit inside the reference-derived boundary extension window outside the HSP-supported piece.",
            "- `[unassigned]`: GFF gene is displayed, but no reference family could be assigned.",
            "",
        ]),
        encoding="utf-8",
    )
    print(f"[content] OUTPUTS.md - {out_path}")


# ── pieces.csv ──────────────────────────────────────────────────────
def write_pieces_csv(
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
) -> None:
    locus_id = locus_cfg["locus_id"]
    headers = [
        "genome_acc", "species", "strain", "assembly_level",
        "locus_id", "status", "piece_idx", "tier",
        "contig", "strand",
        "q_start", "q_end", "q_extent_bp",
        "s_lo", "s_hi", "aligned_bp",
        "identity_pct", "n_covered_genes", "covered_genes",
        "internal_breaks", "constituent_hsps",
        # v1.0 fragmentation annotation (populated for FRAGMENTED pieces)
        "contig_edge_proximal", "edge_distance_bp",
    ]
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for acc in sorted(ruler_results):
            res    = ruler_results[acc]
            pieces = res.get("_pieces") or []
            breaks = res.get("_internal_breaks") or []
            if not pieces:
                continue
            meta   = genome_meta.get(acc, {})
            status = res.get("status", "")
            for i, p in enumerate(pieces):
                qs = int(p["qstart"]); qe = int(p["qend"])
                ss = int(p["sstart"]); se = int(p["send"])
                strand  = "+" if se > ss else "-"
                s_lo, s_hi = min(ss, se), max(ss, se)
                q_lo, q_hi = min(qs, qe), max(qs, qe)
                genes    = p.get("_genes") or []
                p_breaks = breaks[i] if i < len(breaks) else []
                _ep = p.get("_edge_proximal")
                _ed = p.get("_edge_distance")
                w.writerow([
                    acc, meta.get("species", ""), meta.get("strain", ""),
                    meta.get("assembly_level", ""), locus_id, status, i,
                    p.get("_tier", ""), p.get("sseqid", ""), strand,
                    q_lo, q_hi, q_hi - q_lo + 1,
                    s_lo, s_hi, int(p.get("length", 0)),
                    p.get("pident", ""), len(genes), ";".join(genes),
                    len(p_breaks), int(p.get("_constituent_hsp_count", 1)),
                    ("Y" if _ep else ("N" if _ep is False else "")),
                    ("" if _ed is None else _ed),
                ])
    print(f"[content] pieces.csv - {out_path}")


# ── clade_markers.tsv ──────────────────────────────────────────────────────
def write_clade_markers_tsv(
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
    bin_bp: int = 50,
    min_genomes: int = 2,
) -> None:
    """Write recurring internal-break positions across genomes."""
    locus_id = locus_cfg["locus_id"]
    by_pos: dict[tuple[int, Optional[str]], dict[str, dict]] = {}
    for acc, res in ruler_results.items():
        for piece_breaks in (res.get("_internal_breaks") or []):
            for b in piece_breaks:
                q_lo  = int(b.get("q_gap_lo") or 0)
                q_bin = (q_lo // bin_bp) * bin_bp
                ins   = int(b.get("subject_insertion_bp") or 0)
                key   = (q_bin, b.get("ref_gene"))
                slot  = by_pos.setdefault(key, {})
                prev  = slot.get(acc)
                if prev is None or ins > prev["subject_insertion"]:
                    slot[acc] = {"subject_insertion": ins,
                                 "q_gap_bp": int(b.get("q_gap_bp") or 0)}

    rows: list[dict] = []
    for (q_bin, ref_gene), per_genome in by_pos.items():
        if len(per_genome) < min_genomes:
            continue
        species_count: dict[str, int] = {}
        for acc in per_genome:
            sp = genome_meta.get(acc, {}).get("species") or ""
            species_count[sp] = species_count.get(sp, 0) + 1
        ins_values = sorted(v["subject_insertion"] for v in per_genome.values())
        median_ins = ins_values[len(ins_values) // 2]
        rows.append({
            "q_pos_bin":                   q_bin,
            "ref_gene":                    ref_gene or "(intergenic)",
            "n_genomes":                   len(per_genome),
            "median_subject_insertion_bp": median_ins,
            "species_breakdown":           ", ".join(
                f"{sp}:{n}" for sp, n in
                sorted(species_count.items(), key=lambda kv: -kv[1])),
            "genome_accs":                 ";".join(sorted(per_genome.keys())),
        })
    rows.sort(key=lambda r: (-r["n_genomes"], r["q_pos_bin"]))

    with open(out_path, "w", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["locus_id", "q_pos_bin", "ref_gene", "n_genomes",
                    "median_subject_insertion_bp", "species_breakdown",
                    "genome_accs"])
        for r in rows:
            w.writerow([locus_id, r["q_pos_bin"], r["ref_gene"], r["n_genomes"],
                        r["median_subject_insertion_bp"],
                        r["species_breakdown"], r["genome_accs"]])
    print(f"[content] clade_markers.tsv - {out_path}  ({len(rows)} recurring breakpoints)")


# ── marker_matrix.csv ──────────────────────────────────────────────────────
def write_marker_matrix_csv(
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
    bin_bp: int = 50,
    min_genomes: int = 2,
) -> None:
    """Binary genome x recurring-breakpoint matrix (drop-in for phylogeny tools)."""
    per_genome: dict[str, set[tuple[int, Optional[str]]]] = {}
    counts: dict[tuple[int, Optional[str]], int] = {}
    for acc, res in ruler_results.items():
        s: set[tuple[int, Optional[str]]] = set()
        for piece_breaks in (res.get("_internal_breaks") or []):
            for b in piece_breaks or []:
                q_lo  = int(b.get("q_gap_lo") or 0)
                q_bin = (q_lo // bin_bp) * bin_bp
                s.add((q_bin, b.get("ref_gene")))
        per_genome[acc] = s
        for key in s:
            counts[key] = counts.get(key, 0) + 1

    markers = sorted(
        [k for k, n in counts.items() if n >= min_genomes],
        key=lambda k: (k[0], k[1] or ""),
    )
    if not markers:
        print("[content] marker_matrix.csv skipped (no recurring breakpoints)")
        return

    headers = ["genome_acc", "species", "strain", "status"] + [
        f"q{pos}_{gene or 'intergenic'}" for (pos, gene) in markers
    ]
    with open(out_path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for acc in sorted(ruler_results):
            res  = ruler_results[acc]
            meta = genome_meta.get(acc, {})
            row  = [acc, meta.get("species", ""), meta.get("strain", ""),
                    res.get("status", "")]
            present = per_genome.get(acc, set())
            row.extend(1 if m in present else 0 for m in markers)
            w.writerow(row)
    print(f"[content] marker_matrix.csv - {out_path}  "
          f"({len(markers)} markers x {len(ruler_results)} genomes)")


# ── loci.xlsx ──────────────────────────────────────────────────────
def write_loci_xlsx(
    all_locus_genes: dict[str, list[dict]],
    ruler_results: dict[str, dict],
    locus_cfg: dict,
    genome_meta: dict[str, dict],
    out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    locus_id   = locus_cfg["locus_id"]
    auto       = locus_cfg.get("_auto", {})
    n_expected = auto.get("n_genes_ref", 0)
    max_genes  = max((len(gs) for gs in all_locus_genes.values()), default=0)
    max_genes  = max(max_genes, n_expected)

    wb = Workbook()
    ws = wb.active
    ws.title = locus_id[:31]
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    fixed_headers = list(LOCI_FIXED_HEADERS)
    gene_headers = [f"g{i+1:03d}" for i in range(max_genes)]
    all_headers  = fixed_headers + gene_headers

    header_fill = PatternFill("solid", fgColor="2F4F4F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, hdr in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = thin
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    status_color = STATUS_TINT

    # Derived from fixed_headers rather than hardcoded, so reordering them keeps the tint right.
    STATUS_COL = fixed_headers.index("status") + 1

    for row_idx, acc in enumerate(_loci_row_order(all_locus_genes, genome_meta), start=2):
        res    = ruler_results.get(acc, {})
        meta   = genome_meta.get(acc, {})
        genes  = all_locus_genes[acc]
        status = res.get("status", "")
        s_fill = PatternFill("solid", fgColor=status_color.get(status, "FFFFFF"))
        cov    = res.get("coverage")
        cov_s  = f"{cov:.1%}" if cov is not None else ""

        fixed_vals = [
            acc, meta.get("species", ""), meta.get("strain", ""),
            meta.get("assembly_level", ""), status, cov_s,
            res.get("cluster_bp", ""), res.get("genes_found", ""),
            res.get("pseudo_count", ""), res.get("contig", ""),
        ]
        for col_idx, val in enumerate(fixed_vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx,
                           value=str(val) if val != "" else "")
            cell.border    = thin
            cell.alignment = Alignment(horizontal="left", wrap_text=False)
            if col_idx == STATUS_COL:
                cell.fill = s_fill
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = Font(size=9)

        for gene_idx, g in enumerate(genes):
            col_idx  = len(fixed_headers) + gene_idx + 1
            is_ps    = g.get("is_pseudo", False)
            zone     = g.get("_zone")
            label, is_divergent, dna_style, strong, detail = _cell_label_parts(g, res)
            # Decayed-state borders and T/D markers apply to cluster-zone cells only.
            signal_word = ""
                # T+D - no marker (normal), neither - no marker (uninformative)

            color  = zone_color(zone, is_ps, divergent=is_divergent,
                                  dna_only=dna_style)
            cell    = ws.cell(row=row_idx, column=col_idx)
            cell.border    = thin
            cell.alignment = Alignment(horizontal="left", wrap_text=False)
            cell.fill      = PatternFill("solid", fgColor=color)
            if zone == "separator":
                cell.value = g.get("product", "|||")
                cell.font  = Font(bold=True, color="FFFFFF", size=9)
                cell.alignment = Alignment(horizontal="center", wrap_text=False)
            else:
                cell.value = f"{label} {g['product'][:40]}"
                if detail:
                    cell.comment = Comment(detail, "LocusRuler")
                cell.font  = Font(italic=(is_ps or is_divergent or dna_style),
                                  color=("FFFFFF" if strong else "000000"),
                                  bold=strong, size=9)

        for extra in range(len(genes), max_genes):
            col_idx = len(fixed_headers) + extra + 1
            ws.cell(row=row_idx, column=col_idx, value="").border = thin

    for col_idx, name in enumerate(fixed_headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            LOCI_FIXED_WIDTHS.get(name, 14))
    for i in range(max_genes):
        ws.column_dimensions[get_column_letter(len(fixed_headers) + i + 1)].width = 30
    ws.freeze_panes = freeze_after_status(fixed_headers)
    wb.save(out_path)
    print(f"[content] loci.xlsx - {out_path}")
