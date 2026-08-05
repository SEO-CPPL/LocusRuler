"""The tables/ and diagnostics/ split, and the path resolver."""

import tempfile
import unittest
from pathlib import Path

from locus_ruler.writers import (
    DIAGNOSTICS_DIRNAME,
    DIAGNOSTIC_FILES,
    TABLES_DIRNAME,
    TABLE_FILES,
    diagnostics_dir,
    find_output,
    output_home,
    tables_dir,
)


class DiagnosticsDirTests(unittest.TestCase):
    def test_creates_the_subdirectory(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            made = diagnostics_dir(out)
            self.assertTrue(made.is_dir())
            self.assertEqual(made, out / DIAGNOSTICS_DIRNAME)

    def test_is_idempotent(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            first = diagnostics_dir(out)
            (first / "pieces.csv").write_text("x", encoding="utf-8")
            second = diagnostics_dir(out)
            self.assertEqual(first, second)
            self.assertTrue((second / "pieces.csv").exists())


class FindOutputTests(unittest.TestCase):
    def test_prefers_the_diagnostics_copy(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            nested = tables_dir(out) / "pieces.csv"
            nested.write_text("new", encoding="utf-8")
            (out / "pieces.csv").write_text("stale", encoding="utf-8")
            self.assertEqual(find_output(out, "pieces.csv"), nested)
            self.assertEqual(find_output(out, "pieces.csv").read_text(), "new")

    def test_falls_back_to_the_flat_layout(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            flat = out / "gene_diagnostics.csv"
            flat.write_text("legacy", encoding="utf-8")
            self.assertEqual(find_output(out, "gene_diagnostics.csv"), flat)

    def test_missing_file_reports_the_new_location(self):
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            resolved = find_output(out, "pieces.csv")
            self.assertEqual(resolved, out / TABLES_DIRNAME / "pieces.csv")
            self.assertFalse(resolved.exists())

    def test_every_declared_output_file_resolves_both_ways(self):
        for name in TABLE_FILES + DIAGNOSTIC_FILES:
            with self.subTest(filename=name):
                with tempfile.TemporaryDirectory() as tmp:
                    out = Path(tmp)
                    (out / name).write_text("flat", encoding="utf-8")
                    self.assertEqual(find_output(out, name), out / name)
                    nested = Path(out) / output_home(name) / name
                    nested.parent.mkdir(parents=True, exist_ok=True)
                    nested.write_text("nested", encoding="utf-8")
                    self.assertEqual(find_output(out, name), nested)


class ReportTests(unittest.TestCase):
    """The report reads whatever layout is on disk and never invents sheets."""

    def _minimal_run(self, root: Path, nested: bool):
        target = tables_dir(root) if nested else root
        (target / "genome_summary.csv").write_text(
            "genome_acc,species,strain,locus_id,status,coverage\n"
            "GCF_000000001.1,Genus species,A,demo,COMPLETE,96.9%\n"
            "GCF_000000002.1,Genus species,B,demo,DECAYED,42.7%\n",
            encoding="utf-8",
        )
        (target / "pieces.csv").write_text(
            "genome_acc,locus_id,piece_idx,contig,s_lo,s_hi\n"
            "GCF_000000001.1,demo,1,c1,100,900\n",
            encoding="utf-8",
        )

    def test_builds_from_the_nested_layout(self):
        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            out = build_report(root, root / "locus_report.xlsx", "demo")
            self.assertTrue(out.exists())

            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertIn("Overview", wb.sheetnames)
            self.assertIn("Locus results", wb.sheetnames)
            self.assertIn("Pieces", wb.sheetnames)
            # absent sources must not produce empty sheets
            self.assertNotIn("Breakpoints", wb.sheetnames)
            wb.close()

    def test_builds_from_the_legacy_flat_layout(self):
        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=False)
            out = build_report(root, root / "locus_report.xlsx", "demo")
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertIn("Pieces", wb.sheetnames)
            wb.close()

    def test_percent_strings_become_sortable_numbers(self):
        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            out = build_report(root, root / "locus_report.xlsx", "demo")

            import openpyxl
            wb = openpyxl.load_workbook(out)
            ws = wb["Locus results"]
            header = [c.value for c in ws[1]]
            # the sheet projection renames genome_summary's columns
            col = header.index("Coverage") + 1
            values = [ws.cell(row=r, column=col).value for r in (2, 3)]
            self.assertTrue(
                all(isinstance(v, float) for v in values),
                f"coverage must be numeric to sort correctly, got {values!r}",
            )
            self.assertAlmostEqual(max(values), 0.969, places=3)
            self.assertEqual(ws.cell(row=2, column=col).number_format, "0.0%")
            wb.close()

    def test_cassette_matrix_is_folded_in_with_its_formatting(self):
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import PatternFill

        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)

            # stand in for what cassette_structure.py produces
            cassette = tables_dir(root)
            src = openpyxl.Workbook()
            ws = src.active
            ws.title = "cassette_matrix"
            ws.append(["genome_acc", "structure_id", "g001"])
            ws.append(["GCF_000000001.1", "CS_abc", "TAG_0001"])
            tinted = ws.cell(row=2, column=3)
            tinted.fill = PatternFill("solid", fgColor="00FC1B07")
            tinted.comment = Comment("GeneD / INTACT", "locus_ruler")
            ws.column_dimensions["A"].width = 20.0
            ws.freeze_panes = "C2"
            src.save(cassette / "cassette_matrix.xlsx")
            src.close()

            out = build_report(root, root / "locus_report.xlsx", "demo")
            wb = openpyxl.load_workbook(out)
            self.assertIn("Cassette genes", wb.sheetnames)
            got = wb["Cassette genes"]
            cell = got.cell(row=2, column=3)
            self.assertEqual(cell.value, "TAG_0001")
            self.assertEqual(cell.fill.fgColor.rgb, "00FC1B07")
            self.assertIsNotNone(cell.comment)
            self.assertIn("GeneD", cell.comment.text)
            self.assertEqual(got.freeze_panes, "C2")
            self.assertAlmostEqual(got.column_dimensions["A"].width, 20.0)
            wb.close()

    def test_source_workbooks_are_never_modified(self):
        """Step 6 must not write into step 5's file, or `--from 5` would
        silently drop the cassette sheet on a later rerun."""
        import openpyxl

        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            loci = tables_dir(root) / "loci.xlsx"
            src = openpyxl.Workbook()
            src.active.append(["genome_acc", "g001"])
            src.active.append(["GCF_000000001.1", "[GeneE]"])
            src.save(loci)
            src.close()
            before = loci.read_bytes()

            build_report(root, root / "locus_report.xlsx", "demo")
            self.assertEqual(loci.read_bytes(), before,
                             "loci.xlsx must be copied, not appended to")

            wb = openpyxl.load_workbook(loci)
            self.assertEqual(wb.sheetnames, ["Sheet"])
            wb.close()

    def test_both_matrices_land_adjacent(self):
        import openpyxl

        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            for path in (tables_dir(root) / "loci.xlsx",
                         tables_dir(root) / "cassette_matrix.xlsx"):
                path.parent.mkdir(parents=True, exist_ok=True)
                book = openpyxl.Workbook()
                book.active.append(["genome_acc", "g001"])
                book.active.append(["GCF_000000001.1", "x"])
                book.save(path)
                book.close()

            out = build_report(root, root / "locus_report.xlsx", "demo")
            wb = openpyxl.load_workbook(out)
            names = wb.sheetnames
            self.assertIn("Locus genes", names)
            self.assertIn("Cassette genes", names)
            self.assertEqual(
                names.index("Cassette genes") - names.index("Locus genes"), 1,
                f"the two gene grids should be adjacent, got {names}",
            )
            self.assertEqual(names[0], "Overview")
            wb.close()

    def test_overview_lists_sheets_in_tab_order(self):
        import openpyxl

        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            for path in (tables_dir(root) / "loci.xlsx",
                         tables_dir(root) / "cassette_matrix.xlsx"):
                path.parent.mkdir(parents=True, exist_ok=True)
                book = openpyxl.Workbook()
                book.active.append(["genome_acc", "g001"])
                book.active.append(["GCF_000000001.1", "x"])
                book.save(path)
                book.close()

            out = build_report(root, root / "locus_report.xlsx", "demo")
            wb = openpyxl.load_workbook(out)
            tabs = [n for n in wb.sheetnames if n != "Overview"]

            ws = wb["Overview"]
            header_row = next(
                r for r in range(1, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == "Sheets in this workbook"
            )
            listed = []
            for r in range(header_row + 1, ws.max_row + 1):
                value = ws.cell(row=r, column=1).value
                if value in tabs:
                    listed.append(value)
                elif listed:
                    break
            self.assertEqual(listed, [t for t in tabs if t in listed],
                             f"Overview order {listed} != tab order {tabs}")
            wb.close()

    def test_missing_cassette_matrix_is_not_fatal(self):
        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            self._minimal_run(root, nested=True)
            out = build_report(root, root / "locus_report.xlsx", "demo")
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertNotIn("Cassette genes", wb.sheetnames)
            self.assertIn("Locus results", wb.sheetnames)
            wb.close()

    def test_empty_directory_raises_rather_than_writing_a_blank_book(self):
        from locus_ruler.report import build_report

        with tempfile.TemporaryDirectory() as tmp:
            with self.assertRaises(FileNotFoundError):
                build_report(Path(tmp), Path(tmp) / "locus_report.xlsx", "demo")


if __name__ == "__main__":
    unittest.main()
